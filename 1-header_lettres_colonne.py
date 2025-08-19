#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "Sircom_vide_na.xlsx",
# d'ajouter la lettre de chaque colonne suivie d'un underscore (_) au début de chaque en-tête de colonne,
# et d'enregistrer le fichier sous "1-header-lettres-colonne-excel-mapping-excel.xlsx"

# python3 1-header_lettres_colonne.py

import openpyxl
import os

# 1. Définir le fichier source
file_path = "Sircom_vide_na.xlsx"

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"❌ Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print("💡 Assurez-vous d'avoir exécuté le script '0-si-cellule-vide-na.py' au préalable.")
    exit(1)

print(f"📂 Traitement du fichier : {file_path}")

# 3. Ouvrir le fichier Excel
try:
    workbook = openpyxl.load_workbook(file_path)
    print(f"✅ Fichier ouvert avec succès")
except Exception as e:
    print(f"❌ Erreur lors de l'ouverture du fichier : {e}")
    exit(1)

try:
    # 4. Parcourir tous les onglets
    for sheet_name in workbook.sheetnames:
        print(f"🔄 Traitement de la feuille : {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Parcourir les en-têtes de colonnes (première ligne) et ajouter la lettre de colonne
        for column in sheet.iter_cols(min_row=1, max_row=1):
            for cell in column:
                if cell.value is not None:  # Vérifier que la cellule n'est pas vide
                    letter = openpyxl.utils.get_column_letter(cell.column)
                    original_value = str(cell.value)
                    cell.value = f"{letter}_{original_value}"
                    print(f"  📝 Colonne {letter}: '{original_value}' → '{cell.value}'")
    
    # 5. Enregistrer le fichier modifié
    output_filename = "1-header-lettres-colonne-excel-mapping-excel.xlsx"
    workbook.save(output_filename)
    print(f"✅ Fichier sauvegardé sous : {output_filename}")
    print("🎉 Modification terminée avec succès !")

except Exception as e:
    print(f"❌ Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 6. Fermer le fichier Excel
    workbook.close()
    print("📁 Fichier fermé")