#!/usr/bin/env python3

"""
Script 11 - Créer un mapping Excel/CSV entre les colonnes originales et finales
avec indication des champs demandés par Charles

Ce script génère deux fichiers :
- mapping_colonnes_charles.xlsx : avec formatage et surlignage vert
- mapping_colonnes_charles.csv : format simple

Usage : python3 11-create_mapping_excel.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

# Champs demandés par Charles selon Expression-de-besoin.md
CHARLES_REQUESTED_COLUMNS = {
    'E': 'Nom de l\'entreprise',
    'AW': 'Nom du produit',
    'BH': 'Ville de production',
    'AY': 'Catégorie du produit',
    'AX': 'Description du produit',
    'BM': 'Certification OFG',
    'BS': 'Prix départ usine',
    'BQ': '% de valeur ajouté en France',
    'BD': 'Exportation',
    'AD': 'Type d\'entreprise',
    'AE': 'Nombre de salariés',
    'AF': 'Chiffre d\'affaires',
    'AI': 'Présentation de l\'entreprise',
    'AN': 'Label EPV',
    'AU': 'Programme du gouvernement',
    'AJ': 'Démarche de relocalisation',
    'BF': 'Page de présentation régionale',
    'AM': 'Lequel ? (pour logo)'
}

def get_excel_headers(filepath):
    """Lire les en-têtes du fichier Excel après l'étape 1"""
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(cell.value)
    wb.close()
    return headers

def get_csv_headers(filepath):
    """Lire les en-têtes du fichier CSV final"""
    # Le fichier est en UTF-16
    df = pd.read_csv(filepath, encoding='utf-16', nrows=0)
    return list(df.columns)

def create_mapping():
    """Créer le mapping entre les colonnes Excel et CSV"""
    
    print("📊 Création du mapping des colonnes...")
    
    # Lire les en-têtes Excel
    excel_file = "1-header-lettres-colonne-excel-mapping-excel.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Erreur : Le fichier {excel_file} n'existe pas")
        print("💡 Assurez-vous d'avoir exécuté les scripts de traitement")
        return
    
    excel_headers = get_excel_headers(excel_file)
    print(f"✅ {len(excel_headers)} colonnes Excel trouvées")
    
    # Lire les en-têtes CSV
    csv_file = "9-final-sircom-indesign-utf16.csv"
    if not os.path.exists(csv_file):
        print(f"❌ Erreur : Le fichier {csv_file} n'existe pas")
        return
    
    csv_headers = get_csv_headers(csv_file)
    print(f"✅ {len(csv_headers)} colonnes CSV trouvées")
    
    # Créer le mapping
    mapping_data = []
    
    for excel_col in excel_headers:
        # Extraire la lettre de colonne
        col_letter = excel_col.split('_')[0] if '_' in excel_col else ''
        
        # Trouver la colonne CSV correspondante
        csv_col = None
        for csv_h in csv_headers:
            if csv_h.lower().startswith(col_letter.lower() + '_'):
                csv_col = csv_h
                break
        
        # Vérifier si Charles l'a demandé
        requested = 'Oui' if col_letter in CHARLES_REQUESTED_COLUMNS else 'Non'
        
        # Ajouter au mapping
        mapping_data.append({
            'Colonne Excel Original': excel_col,
            'Colonne CSV Final': csv_col if csv_col else 'Non mappé',
            'Demandé par Charles': requested,
            'Description': CHARLES_REQUESTED_COLUMNS.get(col_letter, '')
        })
    
    # Ajouter les colonnes spéciales du CSV
    special_cols = ['imageid', '@pathimg']
    for col in special_cols:
        if col in csv_headers:
            mapping_data.append({
                'Colonne Excel Original': f'(Ajouté) {col}',
                'Colonne CSV Final': col,
                'Demandé par Charles': 'Non',
                'Description': 'Colonne ajoutée pour InDesign'
            })
    
    # Créer le DataFrame
    df = pd.DataFrame(mapping_data)
    
    # Sauvegarder en CSV
    csv_output = "mapping_colonnes_charles.csv"
    df.to_csv(csv_output, index=False, encoding='utf-8-sig')
    print(f"✅ Fichier CSV créé : {csv_output}")
    
    # Sauvegarder en Excel avec formatage
    excel_output = "mapping_colonnes_charles.xlsx"
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Mapping')
        
        # Obtenir la feuille pour le formatage
        worksheet = writer.sheets['Mapping']
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Formater les en-têtes
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Surligner les lignes demandées par Charles
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row[2].value == 'Oui':  # Colonne "Demandé par Charles"
                for cell in row:
                    cell.fill = green_fill
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Fichier Excel créé : {excel_output}")
    
    # Afficher un résumé
    print(f"\n📋 Résumé du mapping :")
    print(f"  - Total de colonnes : {len(mapping_data)}")
    print(f"  - Colonnes demandées par Charles : {len([m for m in mapping_data if m['Demandé par Charles'] == 'Oui'])}")
    print(f"  - Colonnes non mappées : {len([m for m in mapping_data if m['Colonne CSV Final'] == 'Non mappé'])}")
    
    # Afficher les colonnes demandées par Charles
    print(f"\n🎯 Colonnes demandées par Charles :")
    for m in mapping_data:
        if m['Demandé par Charles'] == 'Oui':
            print(f"  - {m['Colonne Excel Original']} → {m['Colonne CSV Final']}")

if __name__ == "__main__":
    create_mapping()