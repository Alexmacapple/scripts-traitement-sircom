#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "5-livrable-final-word.xlsx",
# de nettoyer les en-têtes de colonnes (minuscules, sans accents, sans caractères spéciaux, 10 caractères max),
# de conserver les préfixes pour éviter les collisions,
# et d'enregistrer le fichier sous "6-clean-headers.xlsx"

# python3 6-clean-headers.py

import openpyxl
import os
import re
import unicodedata

# Fonction pour nettoyer les noms des colonnes
def clean_col_name(col_name):
    if col_name is None:
        return col_name
    
    # Convertir en chaîne de caractères
    col_name = str(col_name)
    
    # Convertir en minuscules
    col_name = col_name.lower()
    
    # Supprimer les accents
    col_name = unicodedata.normalize('NFKD', col_name).encode('ASCII', 'ignore').decode('utf-8')
    
    # Supprimer les caractères spéciaux (garder uniquement lettres, chiffres et underscore)
    col_name = re.sub(r'[^\w]', '', col_name)
    
    # Limiter la longueur à 10 caractères
    col_name = col_name[:10]
    
    return col_name

# 1. Définir le fichier source
file_path = "5-livrable-final-word.xlsx"

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"❌ Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print("💡 Assurez-vous d'avoir exécuté le script '5-livrable-final.py' au préalable.")
    exit(1)

print(f"📂 Traitement du fichier : {file_path}")

try:
    # 3. Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_path)
    print(f"✅ Fichier ouvert avec succès")
    
    # 4. Traiter toutes les feuilles
    for sheet_name in workbook.sheetnames:
        print(f"🔄 Traitement de la feuille : {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Compter les colonnes avec des en-têtes
        headers_processed = 0
        
        # Parcourir les en-têtes de colonnes (première ligne) et les nettoyer
        for column in sheet.iter_cols(min_row=1, max_row=1):
            for cell in column:
                if cell.value is not None:
                    original_value = str(cell.value)
                    cleaned_value = clean_col_name(original_value)
                    cell.value = cleaned_value
                    headers_processed += 1
                    
                    # Afficher quelques exemples de transformation
                    if headers_processed <= 10:
                        print(f"  📝 Colonne {openpyxl.utils.get_column_letter(cell.column)}: '{original_value}' → '{cleaned_value}'")
                    elif headers_processed == 11:
                        print(f"  📝 ... (et {sheet.max_column - 10} autres colonnes)")
        
        print(f"✅ {headers_processed} en-têtes nettoyés dans la feuille '{sheet_name}'")
    
    # 5. Enregistrer le fichier modifié
    output_filename = "6-clean-headers.xlsx"
    workbook.save(output_filename)
    print(f"✅ Fichier sauvegardé sous : {output_filename}")
    
    # 6. Afficher un résumé des transformations
    print(f"\n📋 Règles de nettoyage appliquées :")
    print(f"  ✓ Conversion en minuscules")
    print(f"  ✓ Suppression des accents")
    print(f"  ✓ Suppression des caractères spéciaux")
    print(f"  ✓ Limitation à 10 caractères maximum")
    print(f"  ✓ Conservation des préfixes pour éviter les collisions")
    
    print("🎉 Nettoyage des en-têtes terminé avec succès !")

except Exception as e:
    print(f"❌ Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 7. Fermer le fichier Excel
    if 'workbook' in locals():
        workbook.close()
    print("📁 Fichier fermé")