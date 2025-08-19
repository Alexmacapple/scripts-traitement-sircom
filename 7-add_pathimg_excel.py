#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "6-clean-headers.xlsx",
# d'ajouter une colonne "@pathimg" juste après la colonne "imageid",
# de générer les chemins d'images au format "dossier-{ID}.jpg" pour InDesign,
# et d'enregistrer le fichier sous "7-add-pathimg.xlsx"

# python3 7-add-pathimg.py

import openpyxl
import os

# Configuration du chemin des images (format POSIX pour InDesign 19.4+)
IMAGE_BASE_PATH = "/Users/victoria/Documents/export-jpg-resize"

# 1. Définir le fichier source
file_path = "6-clean-headers.xlsx"

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"❌ Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print("💡 Assurez-vous d'avoir exécuté le script '6-clean-headers.py' au préalable.")
    exit(1)

print(f"📂 Traitement du fichier : {file_path}")
print(f"🖼️  Chemin des images configuré : {IMAGE_BASE_PATH}")

try:
    # 3. Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_path)
    print(f"✅ Fichier ouvert avec succès")
    
    # 4. Traiter la feuille active
    worksheet = workbook.active
    sheet_name = worksheet.title
    print(f"🔄 Traitement de la feuille : {sheet_name}")

    # 5. Chercher la colonne "imageid"
    imageid_column = None
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == "imageid":
                imageid_column = cell.column
                print(f"✅ Colonne 'imageid' trouvée en position : {openpyxl.utils.get_column_letter(imageid_column)}")
                break
        if imageid_column:
            break

    if imageid_column:
        # 6. Insérer une nouvelle colonne "@pathimg" juste après "imageid"
        pathimg_column = imageid_column + 1
        worksheet.insert_cols(pathimg_column)
        worksheet.cell(row=1, column=pathimg_column).value = "@pathimg"
        print(f"✅ Colonne '@pathimg' ajoutée en position : {openpyxl.utils.get_column_letter(pathimg_column)}")

        # 7. Parcourir les données et générer les chemins d'images
        rows_processed = 0
        paths_generated = 0
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            rows_processed += 1
            
            # Récupérer la valeur de imageid (décalée d'une position à cause de l'insertion)
            imageid_cell = row[imageid_column - 1]  # -1 car l'index de la liste commence à 0
            
            if imageid_cell.value is not None and imageid_cell.value != "#N/A":
                # Extraire l'ID depuis le nom de fichier "dossier-XXXXX.jpg"
                imageid_value = str(imageid_cell.value)
                
                # Si c'est déjà au format "dossier-ID.jpg", extraire l'ID
                if imageid_value.startswith("dossier-") and imageid_value.endswith(".jpg"):
                    # Extraire l'ID : "dossier-24331205.jpg" → "24331205"
                    image_id = imageid_value.replace("dossier-", "").replace(".jpg", "")
                else:
                    # Sinon, utiliser la valeur telle quelle comme ID
                    image_id = imageid_value
                
                # Générer le chemin complet (format POSIX)
                full_path = f"{IMAGE_BASE_PATH}/dossier-{image_id}.jpg"
                worksheet.cell(row=imageid_cell.row, column=pathimg_column).value = full_path
                paths_generated += 1
                
                # Afficher les 5 premiers exemples
                if paths_generated <= 5:
                    print(f"  📝 Ligne {imageid_cell.row}: ID={image_id} → {full_path}")
            else:
                # Pour les cellules vides, mettre #N/A
                worksheet.cell(row=imageid_cell.row, column=pathimg_column).value = "#N/A"
        
        print(f"✅ {paths_generated}/{rows_processed} chemins d'images générés")

        # 8. Enregistrer les modifications
        output_filename = "7-add-pathimg.xlsx"
        workbook.save(output_filename)
        print(f"✅ Fichier sauvegardé sous : {output_filename}")
        
        # 9. Afficher un résumé
        print(f"\n📋 Résumé du traitement :")
        print(f"  ✓ Colonne source : 'imageid'")
        print(f"  ✓ Colonne ajoutée : '@pathimg'")
        print(f"  ✓ Format des chemins : {IMAGE_BASE_PATH}/dossier-{{ID}}.jpg")
        print(f"  ✓ Lignes traitées : {rows_processed}")
        print(f"  ✓ Chemins générés : {paths_generated}")
        
        print("🎉 Ajout des chemins d'images terminé avec succès !")
        
    else:
        print("❌ Erreur : La colonne 'imageid' n'a pas été trouvée dans la feuille.")
        print("💡 Colonnes disponibles :")
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value:
                    print(f"  - {cell.value}")
        exit(1)

except Exception as e:
    print(f"❌ Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 10. Fermer le fichier Excel
    if 'workbook' in locals():
        workbook.close()
    print("📁 Fichier fermé")