#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script d'extraction d'un onglet Excel vers 00-sircom-source.xlsx
======================================================

Ce script extrait par défaut l'onglet 2026 "BDD TT + ANALYSE DGDDI"
du fichier Excel multi-onglets et le sauvegarde comme fichier 00-sircom-source.xlsx
pour traitement ultérieur avec la copie des anciens scripts.

Auteur : Claude Assistant
Date : 17 septembre 2025
"""

import sys
import argparse
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from sircom2026_rules import DEFAULT_SHEET_NAME, config_value, resolve_repo_path


DEFAULT_SOURCE_FILE = str(resolve_repo_path(config_value("excel_source_path")))
DEFAULT_OUTPUT_FILE = config_value("step_00_output")
DEFAULT_SOURCE_SHEET = config_value("sheet_name", DEFAULT_SHEET_NAME)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extraire un onglet vers 00-sircom-source.xlsx.",
    )
    parser.add_argument(
        "--source",
        default=DEFAULT_SOURCE_FILE,
        help=f"Fichier Excel source (défaut : {DEFAULT_SOURCE_FILE}).",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Fichier Excel de sortie (défaut : {DEFAULT_OUTPUT_FILE}).",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SOURCE_SHEET,
        help=f"Onglet source à extraire (défaut : {DEFAULT_SOURCE_SHEET}).",
    )
    parser.add_argument(
        "--include-hidden-rows",
        action="store_true",
        help="Conserver les lignes masquées au lieu de les ignorer.",
    )
    return parser.parse_args()


def extract_departements_to_sircom(
    source_file=DEFAULT_SOURCE_FILE,
    output_file=DEFAULT_OUTPUT_FILE,
    sheet_name=DEFAULT_SOURCE_SHEET,
    include_hidden_rows=False,
):
    """
    Extrait l'onglet demandé et crée le fichier 00-sircom-source.xlsx.
    """

    source_path = Path(source_file)
    output_path = Path(output_file)

    print("=" * 60)
    print("EXTRACTION D'UN ONGLET VERS SIRCOM.XLSX")
    print("=" * 60)
    print()

    # Vérifier que le fichier source existe
    if not source_path.exists():
        print("ERREUR : Le fichier source n'existe pas :")
        print(f"   {source_path}")
        sys.exit(1)

    print(f"Fichier source : {source_path.name}")

    try:
        # Lire l'onglet demandé
        print(f"\nLecture de l'onglet '{sheet_name}'...")
        source_workbook = load_workbook(source_path, data_only=True)
        if sheet_name not in source_workbook.sheetnames:
            raise ValueError(f"Onglet '{sheet_name}' introuvable.")
        source_sheet = source_workbook[sheet_name]
        rows = []
        hidden_rows_skipped = 0
        for row in source_sheet.iter_rows():
            row_number = row[0].row
            if (
                row_number > 1
                and not include_hidden_rows
                and source_sheet.row_dimensions[row_number].hidden
            ):
                hidden_rows_skipped += 1
                continue
            rows.append([cell_display_value(cell) for cell in row])
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # Afficher les informations sur les données
        print("\nOnglet lu avec succès :")
        print(f"   - Nombre de lignes : {len(data_rows)}")
        print(f"   - Nombre de colonnes : {len(headers)}")
        print(f"   - Lignes masquées ignorées : {hidden_rows_skipped}")

        # Afficher un aperçu des colonnes
        print("\nAperçu des colonnes :")
        for i, col in enumerate(headers[:5]):
            print(f"   - Colonne {i}: {col}")
        if len(headers) > 5:
            print(f"   ... ({len(headers) - 5} colonnes supplémentaires)")

        # Sauvegarder dans le nouveau fichier
        print(f"\nSauvegarde vers : {output_path.name}")

        # Si le fichier existe déjà, le sauvegarder
        if output_path.exists():
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_path = output_path.with_name(
                f"{output_path.stem}_backup_{timestamp}{output_path.suffix}"
            )
            print(
                f"Le fichier existe déjà, création d'une sauvegarde : {backup_path.name}"
            )
            output_path.rename(backup_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = sheet_name[:31]
        for row in rows:
            output_sheet.append(row)
        output_workbook.save(output_path)
        output_workbook.close()
        source_workbook.close()

        print("\nExtraction terminée avec succès !")
        print("   Le fichier 00-sircom-source.xlsx est prêt pour le traitement.")

        # Vérification finale
        verif_workbook = load_workbook(output_path, data_only=True)
        verif_sheet = verif_workbook[sheet_name[:31]]
        print("\nVérification du fichier créé :")
        print(f"   - Lignes : {max(verif_sheet.max_row - 1, 0)}")
        print(f"   - Colonnes : {verif_sheet.max_column}")

        if max(verif_sheet.max_row - 1, 0) == len(
            data_rows
        ) and verif_sheet.max_column == len(headers):
            print("   Intégrité des données confirmée")
        else:
            print("   Attention : différence détectée dans les dimensions")
        verif_workbook.close()

    except Exception as e:
        print(f"\nERREUR lors de l'extraction : {e}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("Prochaine étape :")
    print("   Lancer ensuite 01-si-cellule-vide-na.py")
    print("   puis lancer : python3 sircom_master_script.py --verbose")
    print("=" * 60)


def cell_display_value(cell):
    value = cell.value
    if value is None:
        return None
    if isinstance(value, (int, float)) and float(value).is_integer():
        number_format = str(cell.number_format or "")
        if re.fullmatch(r"0+", number_format):
            return str(int(value)).zfill(len(number_format))
    return value


if __name__ == "__main__":
    args = parse_args()
    extract_departements_to_sircom(
        args.source,
        args.output,
        args.sheet,
        include_hidden_rows=args.include_hidden_rows,
    )
