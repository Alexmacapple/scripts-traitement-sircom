#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "02-header-lettres-colonne.xlsx",
# d'ajouter une colonne "imageid" juste après la colonne ID dossier,
# de générer des noms de fichiers basés sur Dossier ID (format: "[ID].jpg"),
# et d'enregistrer le fichier sous "03-image-id.xlsx"

# python3 03-image_id_adder.py

import os

import openpyxl

from sircom2026_rules import (
    config_value,
    find_dossier_id_column,
    image_id_for_dossier,
    imageid_column_name,
    is_empty_value,
)

# 1. Définir le fichier source
file_path = config_value("step_02_output")

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print(
        "Assurez-vous d'avoir exécuté le script '02-header_lettres_colonne.py' au préalable."
    )
    exit(1)

print(f"Traitement du fichier : {file_path}")

# 3. Ouvrir le fichier Excel
try:
    workbook = openpyxl.load_workbook(file_path)
    print("Fichier ouvert avec succès")
except Exception as e:
    print(f"Erreur lors de l'ouverture du fichier : {e}")
    exit(1)

try:
    # 4. Sélectionner la première feuille (peu importe son nom)
    worksheet = workbook.active
    sheet_name = worksheet.title
    print(f"Traitement de la feuille : {sheet_name}")

    # 5. Chercher la colonne ID dossier, quelle que soit sa lettre 2025/2026
    id_column, id_header = find_dossier_id_column(worksheet)
    if id_column:
        print(
            f"Colonne ID trouvée ({id_header}) en position : "
            f"{openpyxl.utils.get_column_letter(id_column)}"
        )

    if id_column:
        # 6. Insérer une nouvelle colonne "imageid" juste après l'ID dossier
        image_id_column = id_column + 1
        worksheet.insert_cols(image_id_column)
        image_header = imageid_column_name()
        worksheet.cell(row=1, column=image_id_column).value = image_header
        print(
            f"Colonne '{image_header}' ajoutée en position : {openpyxl.utils.get_column_letter(image_id_column)}"
        )

        # 7. Parcourir les données de la colonne ID et générer imageid
        rows_processed = 0
        imageids_seen = {}
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            cell = row[id_column - 1]
            if not is_empty_value(cell.value):
                file_name = image_id_for_dossier(cell.value)
                if file_name in imageids_seen:
                    raise ValueError(
                        f"Collision imageid {file_name} aux lignes "
                        f"{imageids_seen[file_name]} et {cell.row}"
                    )
                imageids_seen[file_name] = cell.row
                worksheet.cell(row=cell.row, column=image_id_column).value = file_name
                rows_processed += 1
                if rows_processed <= 5:  # Afficher les 5 premiers exemples
                    print(f"  Ligne {cell.row}: ID={cell.value} → {file_name}")

        print(f"{rows_processed} noms de fichiers générés")

        # 8. Enregistrer les modifications dans le fichier Excel
        output_filename = config_value("step_03_output")
        workbook.save(output_filename)
        print(f"Fichier sauvegardé sous : {output_filename}")
        print("Modification terminée avec succès !")

    else:
        print("Erreur : aucune colonne ID dossier n'a été trouvée dans la feuille.")
        print("Vérifiez que le script précédent a bien été exécuté.")
        exit(1)

except Exception as e:
    print(f"Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 9. Fermer le fichier Excel
    workbook.close()
    print("Fichier fermé")
