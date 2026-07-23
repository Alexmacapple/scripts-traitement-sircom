#!/usr/bin/env python3

### Ce script permet :

# Trie le fichier par région puis département quand les colonnes sont présentes.

import os
import re
import unicodedata

import openpyxl

from sircom2026_rules import config_list, config_value

# 1. Définir les fichiers source et destination
input_file = config_value("step_03_output")
output_file = config_value("step_04_output")


def normalize_header(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("utf-8")
    return re.sub(r"[^a-z0-9]", "", text.lower())


def find_column_indexes(sheet):
    headers = [cell.value for cell in sheet[1]]
    normalized = {
        normalize_header(header): index for index, header in enumerate(headers, start=1)
    }
    region_index = find_region_header(normalized)
    departement_index = find_departement_header(normalized)
    return region_index, departement_index


def find_region_header(normalized_headers):
    expected_fragments = config_list("sort_region_header_contains", default="region")
    for header, index in normalized_headers.items():
        if any(fragment in header for fragment in expected_fragments):
            return index
    return None


def find_departement_header(normalized_headers):
    expected_fragments = config_list(
        "sort_departement_header_contains",
        default="departement",
    )
    excluded_fragments = config_list(
        "sort_departement_header_excludes",
        default="postal",
    )
    for header, index in normalized_headers.items():
        if any(fragment in header for fragment in expected_fragments) and not any(
            fragment in header for fragment in excluded_fragments
        ):
            return index
    return None


def department_sort_key(value):
    text = "" if value is None else str(value).strip()
    match = re.match(r"^(\d{1,3})", text)
    if match:
        return (0, int(match.group(1)), text)
    return (1, text.casefold())


def sort_region_departement(sheet):
    region_index, departement_index = find_column_indexes(sheet)
    if region_index is None or departement_index is None:
        return False

    rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    rows.sort(
        key=lambda row: (
            ""
            if row[region_index - 1] is None
            else str(row[region_index - 1]).casefold(),
            department_sort_key(row[departement_index - 1]),
        )
    )
    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    return True


# 2. Vérifier que le fichier source existe
if not os.path.exists(input_file):
    print(
        f"Erreur : Le fichier '{input_file}' n'existe pas dans le répertoire courant."
    )
    print("Assurez-vous d'avoir exécuté le script '03-image_id_adder.py' au préalable.")
    exit(1)

print(f"Traitement du fichier : {input_file}")

try:
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    sorted_rows = sort_region_departement(sheet)
    workbook.save(output_file)
    workbook.close()
    print(f"Fichier sauvegardé : {output_file}")
    if sorted_rows:
        print("Tri région/département appliqué sur la feuille active.")
    else:
        print("Colonnes Région/Département introuvables : ordre source conservé.")
    print("Opération terminée avec succès !")
except Exception as e:
    print(f"Erreur lors du tri région/département : {e}")
    exit(1)
