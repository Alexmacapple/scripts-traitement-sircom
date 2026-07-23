#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "07-clean-headers.xlsx",
# d'ajouter une colonne "@pathimg" juste après la colonne "imageid",
# de générer les chemins d'images au format "{ID}.jpg" pour InDesign,
# et d'enregistrer le fichier sous "08-add-pathimg.xlsx"

# python3 08-add_pathimg_excel.py

import os
from collections import defaultdict

import openpyxl

from sircom2026_rules import (
    DEFAULT_IMAGE_BASE_PATH,
    config_value,
    display_text,
    empty_cell_marker,
    image_id_for_dossier,
    imageid_column_name,
    is_empty_value,
    pathimg_column_name,
    pathimg_path,
)


# Configuration du chemin des images InDesign
IMAGE_BASE_PATH = config_value("pathimg_root", DEFAULT_IMAGE_BASE_PATH)
IMAGEID_HEADER = imageid_column_name()
PATHIMG_HEADER = pathimg_column_name()

# 1. Définir le fichier source
file_path = config_value("step_07_output")

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print(
        "Assurez-vous d'avoir exécuté le script '07-clean_headers_excel.py' au préalable."
    )
    exit(1)

print(f"Traitement du fichier : {file_path}")
print(f"Chemin des images configuré : {IMAGE_BASE_PATH}")

try:
    # 3. Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_path)
    print("Fichier ouvert avec succès")

    # 4. Traiter la feuille active
    worksheet = workbook.active
    sheet_name = worksheet.title
    print(f"Traitement de la feuille : {sheet_name}")

    # 5. Chercher la colonne "imageid"
    imageid_column = None
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == IMAGEID_HEADER:
                imageid_column = cell.column
                print(
                    f"Colonne '{IMAGEID_HEADER}' trouvée en position : {openpyxl.utils.get_column_letter(imageid_column)}"
                )
                break
        if imageid_column:
            break

    if imageid_column:
        # 6. Insérer une nouvelle colonne "@pathimg" juste après "imageid"
        pathimg_column = imageid_column + 1
        worksheet.insert_cols(pathimg_column)
        worksheet.cell(row=1, column=pathimg_column).value = PATHIMG_HEADER
        print(
            f"Colonne '{PATHIMG_HEADER}' ajoutée en position : {openpyxl.utils.get_column_letter(pathimg_column)}"
        )

        # 7. Parcourir les données et générer les chemins d'images
        rows_processed = 0
        paths_generated = 0
        normalized_ids_by_imageid = defaultdict(list)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            rows_processed += 1

            # Récupérer la valeur de imageid (décalée d'une position à cause de l'insertion)
            imageid_cell = row[
                imageid_column - 1
            ]  # -1 car l'index de la liste commence à 0

            if not is_empty_value(imageid_cell.value):
                imageid_value = display_text(imageid_cell.value)
                if imageid_value.lower().endswith((".jpg", ".jpeg", ".png")):
                    normalized_name = image_id_for_dossier(
                        imageid_value.rsplit(".", 1)[0]
                    )
                else:
                    normalized_name = image_id_for_dossier(imageid_value)
                normalized_ids_by_imageid[normalized_name].append(imageid_cell.row)
                if len(normalized_ids_by_imageid[normalized_name]) > 1:
                    rows = ", ".join(
                        str(row_number)
                        for row_number in normalized_ids_by_imageid[normalized_name]
                    )
                    raise ValueError(
                        f"Collision imageid {normalized_name} aux lignes {rows}"
                    )

                # Générer le chemin complet avec le séparateur adapté au chemin racine.
                full_path = pathimg_path(IMAGE_BASE_PATH, normalized_name)

                # Aussi mettre à jour la colonne imageid avec le nom normalisé
                worksheet.cell(
                    row=imageid_cell.row, column=imageid_column
                ).value = normalized_name
                worksheet.cell(
                    row=imageid_cell.row, column=pathimg_column
                ).value = full_path
                paths_generated += 1

                # Afficher les 5 premiers exemples
                if paths_generated <= 5:
                    print(
                        f"  Ligne {imageid_cell.row}: imageid={normalized_name} → {full_path}"
                    )
            else:
                worksheet.cell(
                    row=imageid_cell.row,
                    column=pathimg_column,
                ).value = empty_cell_marker()

        print(f"{paths_generated}/{rows_processed} chemins d'images générés")

        # 8. Enregistrer les modifications
        output_filename = config_value("step_08_output")
        workbook.save(output_filename)
        print(f"Fichier sauvegardé sous : {output_filename}")

        # 9. Afficher un résumé
        print("\nRésumé du traitement :")
        print(f"  Colonne source : '{IMAGEID_HEADER}'")
        print(f"  Colonne ajoutée : '{PATHIMG_HEADER}'")
        print(f"  Format des chemins : {pathimg_path(IMAGE_BASE_PATH, '{ID}.jpg')}")
        print(f"  Lignes traitées : {rows_processed}")
        print(f"  Chemins générés : {paths_generated}")

        print("Ajout des chemins d'images terminé avec succès !")

    else:
        print(
            f"Erreur : La colonne '{IMAGEID_HEADER}' n'a pas été trouvée dans la feuille."
        )
        print("Colonnes disponibles :")
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value:
                    print(f"  - {cell.value}")
        exit(1)

except Exception as e:
    print(f"Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 10. Fermer le fichier Excel
    if "workbook" in locals():
        workbook.close()
    print("Fichier fermé")
