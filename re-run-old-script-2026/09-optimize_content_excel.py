#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "08-add-pathimg.xlsx",
# d'optimiser le contenu pour InDesign (nettoyage espaces, \n→<br>, suppression colonnes/lignes entièrement vides),
# de conserver les lignes partiellement remplies (avec des #N/A),
# et d'enregistrer le fichier sous "09-optimize-content.xlsx"

# python3 09-optimize_content_excel.py

import openpyxl
import os
import re

from sircom2026_rules import (
    config_bool,
    config_value,
    empty_cell_marker,
    is_dossier_id_header,
)


# Fonction pour vérifier si une cellule est considérée comme vide
def is_empty_cell(value):
    # Cas simples
    if value is None:
        return True

    # Convertir en chaîne pour vérifier le contenu
    str_value = str(value).strip()

    # Considérer comme vide si :
    marker = empty_cell_marker()
    return (
        str_value == ""
        or str_value == marker
        or str_value.lower() == "none"
        or str_value.lower() == "undefined"
    )


# Fonction pour nettoyer le contenu d'une cellule
def clean_cell_content(value):
    marker = empty_cell_marker()
    if value is None:
        return marker

    # Convertir en chaîne
    content = str(value)

    # Supprimer les espaces en début/fin
    content = content.strip()

    # Si vide après trim, utiliser #N/A pour éviter les cellules vides InDesign.
    if content == "":
        return marker

    # Remplacer les sauts de ligne par le marqueur InDesign recherché ensuite.
    content = content.replace("\r\n", "\n")
    content = content.replace("\r", "\n")
    content = content.replace("\n", config_value("linebreak_replacement", "<br>"))

    # Supprimer les espaces multiples (remplacer par un seul espace)
    content = re.sub(r" +", " ", content)

    return content


# 1. Définir le fichier source
file_path = config_value("step_08_output")

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print(
        "Assurez-vous d'avoir exécuté le script '08-add_pathimg_excel.py' au préalable."
    )
    exit(1)

print(f"Traitement du fichier : {file_path}")

try:
    # 3. Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_path)
    print("Fichier ouvert avec succès")

    # 4. Traiter la feuille active
    worksheet = workbook.active
    sheet_name = worksheet.title
    print(f"Traitement de la feuille : {sheet_name}")

    original_rows = worksheet.max_row
    original_cols = worksheet.max_column
    print(f"Dimensions initiales : {original_rows} lignes × {original_cols} colonnes")

    # 5. ÉTAPE 1 : Nettoyer le contenu de toutes les cellules
    print("\nÉtape 1 : Nettoyage du contenu des cellules...")
    cells_cleaned = 0

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            if cell.value is not None:
                original_value = cell.value
                cleaned_value = clean_cell_content(original_value)

                if str(original_value) != str(cleaned_value):
                    cell.value = cleaned_value
                    cells_cleaned += 1

    print(f"{cells_cleaned} cellules nettoyées")

    # 6. ÉTAPE 2 : Identifier et supprimer les colonnes entièrement vides
    print("\nÉtape 2 : Suppression des colonnes entièrement vides...")

    columns_to_delete = []

    if config_bool("drop_empty_columns", default=True):
        columns_range = range(1, worksheet.max_column + 1)
    else:
        columns_range = range(0)

    for col_num in columns_range:
        # Vérifier si toute la colonne est vide (sauf l'en-tête qu'on garde toujours)
        column_empty = True

        for row_num in range(
            2, worksheet.max_row + 1
        ):  # Commence à la ligne 2 (garde l'en-tête)
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if not is_empty_cell(cell_value):
                column_empty = False
                break

        if column_empty:
            column_letter = openpyxl.utils.get_column_letter(col_num)
            header_value = worksheet.cell(row=1, column=col_num).value
            columns_to_delete.append((col_num, column_letter, header_value))

    # Supprimer les colonnes (en ordre inverse pour éviter les décalages)
    columns_deleted = 0
    for col_num, col_letter, header in reversed(columns_to_delete):
        print(f"  Suppression colonne {col_letter}: '{header}'")
        worksheet.delete_cols(col_num)
        columns_deleted += 1

    print(f"{columns_deleted} colonnes entièrement vides supprimées")

    # 7. ÉTAPE 3 : Identifier et supprimer les lignes inutiles
    print("\nÉtape 3 : Suppression des lignes inutiles...")

    # Identifier les colonnes critiques
    critical_columns = {}
    for col_num in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=col_num).value
        if header_value:
            header_clean = str(header_value).lower().strip()
            if header_clean in [
                "f_id",
                "id_dossier",
                "imageid",
            ] or is_dossier_id_header(header_value):
                critical_name = (
                    "id_dossier" if is_dossier_id_header(header_value) else header_clean
                )
                critical_columns[critical_name] = col_num
                print(
                    f"  Colonne critique identifiée : '{header_value}' en position {col_num}"
                )

    rows_to_delete = []

    for row_num in range(
        2, worksheet.max_row + 1
    ):  # Commence à la ligne 2 (garde l'en-tête)
        # Vérifier si la ligne doit être supprimée
        should_delete = False
        delete_reason = ""

        # Critère 1 : Ligne entièrement vide
        row_entirely_empty = True
        for col_num in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if not is_empty_cell(cell_value):
                row_entirely_empty = False
                break

        if row_entirely_empty and config_bool("drop_empty_rows", default=True):
            should_delete = True
            delete_reason = "entièrement vide"

        # Critère 2 : Colonnes critiques vides (ID manquant = dossier invalide)
        elif "id_dossier" in critical_columns and config_bool(
            "drop_rows_without_dossier_id", default=True
        ):
            id_value = worksheet.cell(
                row=row_num, column=critical_columns["id_dossier"]
            ).value
            if is_empty_cell(id_value):
                should_delete = True
                delete_reason = "ID manquant (id_dossier vide)"

        if should_delete:
            rows_to_delete.append((row_num, delete_reason))

    # Supprimer les lignes (en ordre inverse pour éviter les décalages)
    rows_deleted = 0
    entirely_empty_deleted = 0
    missing_id_deleted = 0

    for row_num, reason in reversed(rows_to_delete):
        print(f"  Suppression ligne {row_num} ({reason})")
        worksheet.delete_rows(row_num)
        rows_deleted += 1

        if "entièrement vide" in reason:
            entirely_empty_deleted += 1
        elif "ID manquant" in reason:
            missing_id_deleted += 1

    print(f"{rows_deleted} lignes supprimées :")
    print(f"  - {entirely_empty_deleted} lignes entièrement vides")
    print(f"  - {missing_id_deleted} lignes sans ID (dossiers invalides)")

    # 8. Statistiques finales
    final_rows = worksheet.max_row
    final_cols = worksheet.max_column

    print(f"\nDimensions finales : {final_rows} lignes × {final_cols} colonnes")
    print(
        f"Réduction : {original_rows - final_rows} lignes, {original_cols - final_cols} colonnes"
    )

    # 9. Enregistrer le fichier optimisé
    output_filename = config_value("step_09_output")
    workbook.save(output_filename)
    print(f"Fichier sauvegardé sous : {output_filename}")

    # 10. Résumé des optimisations
    print("\nRésumé des optimisations appliquées :")
    print("  Suppression des espaces en début/fin de cellules")
    print("  Remplacement des sauts de ligne par <br>")
    print("  Suppression des espaces multiples")
    print(f"  Remplacement des cellules vides par {empty_cell_marker()}")
    print(f"  Suppression de {columns_deleted} colonnes entièrement vides")
    print(f"  Suppression de {entirely_empty_deleted} lignes entièrement vides")
    print(f"  Suppression de {missing_id_deleted} lignes sans ID (dossiers invalides)")
    print("  Conservation des lignes partiellement remplies avec ID valide")

    print("Optimisation pour InDesign terminée avec succès !")

except Exception as e:
    print(f"Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 11. Fermer le fichier Excel
    if "workbook" in locals():
        workbook.close()
    print("Fichier fermé")
