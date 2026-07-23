#!/usr/bin/env python3

"""
Script 12 - Créer un mapping Excel/CSV entre les colonnes originales et finales
avec indication des champs attendus pour InDesign

Ce script génère deux fichiers :
- 12-mapping-colonnes-sircom-2026.xlsx : avec formatage et surlignage vert
- 12-mapping-colonnes-sircom-2026.csv : format simple

Usage : python3 12-create_mapping_excel.py
"""

import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

from sircom2026_rules import (
    config_list,
    config_mapping,
    config_value,
    is_dossier_id_header,
)

# Champs attendus selon EDB.md - Nouvelle structure 25 colonnes
EXPECTED_MAPPING_COLUMNS = config_mapping("mapping_expected_columns")


def get_excel_headers(filepath):
    """Lire les en-têtes du fichier Excel après l'étape 1"""
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(cell.value)
    wb.close()
    return headers


def get_csv_headers(filepath):
    """Lire les en-têtes du fichier CSV final"""
    with open(
        filepath,
        "r",
        encoding=config_value("csv_encoding", "utf-16"),
        newline="",
    ) as csv_file:
        reader = csv.reader(csv_file)
        return next(reader)


def create_mapping():
    """Créer le mapping entre les colonnes Excel et CSV"""

    print("Création du mapping des colonnes...")

    # Lire les en-têtes Excel
    excel_file = config_value("step_02_output")
    if not os.path.exists(excel_file):
        print(f"Erreur : Le fichier {excel_file} n'existe pas")
        print("Assurez-vous d'avoir exécuté les scripts de traitement")
        return

    excel_headers = get_excel_headers(excel_file)
    print(f"{len(excel_headers)} colonnes Excel trouvées")

    # Lire les en-têtes CSV
    csv_file = config_value("step_10_output")
    if not os.path.exists(csv_file):
        print(f"Erreur : Le fichier {csv_file} n'existe pas")
        return

    csv_headers = get_csv_headers(csv_file)
    print(f"{len(csv_headers)} colonnes CSV trouvées")

    # Créer le mapping
    mapping_data = []

    for excel_col in excel_headers:
        # Extraire la lettre de colonne
        col_letter = excel_col.split("_")[0] if "_" in excel_col else ""

        # Trouver la colonne CSV correspondante
        csv_col = None
        if is_dossier_id_header(excel_col) and "id_dossier" in csv_headers:
            csv_col = "id_dossier"
        else:
            for csv_h in csv_headers:
                if csv_h.lower().startswith(col_letter.lower() + "_"):
                    csv_col = csv_h
                    break

        # Vérifier si la colonne est attendue dans le mapping InDesign
        expected = "Oui" if col_letter in EXPECTED_MAPPING_COLUMNS else "Non"

        # Ajouter au mapping
        mapping_data.append(
            {
                "Colonne Excel Original": excel_col,
                "Colonne CSV Final": csv_col if csv_col else "Non mappé",
                "Champ attendu": expected,
                "Description": EXPECTED_MAPPING_COLUMNS.get(col_letter, ""),
            }
        )

    # Ajouter les colonnes spéciales du CSV
    special_cols = config_list("mapping_special_columns", default="imageid,@pathimg")
    for col in special_cols:
        if col in csv_headers:
            mapping_data.append(
                {
                    "Colonne Excel Original": f"(Ajouté) {col}",
                    "Colonne CSV Final": col,
                    "Champ attendu": "Non",
                    "Description": "Colonne ajoutée pour InDesign",
                }
            )

    # Sauvegarder en CSV
    csv_output = config_value("mapping_csv_output")
    fieldnames = [
        "Colonne Excel Original",
        "Colonne CSV Final",
        "Champ attendu",
        "Description",
    ]
    with open(
        csv_output,
        "w",
        encoding=config_value("mapping_csv_encoding", "utf-8-sig"),
        newline="",
    ) as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(mapping_data)
    print(f"Fichier CSV créé : {csv_output}")

    # Sauvegarder en Excel avec formatage
    excel_output = config_value("mapping_excel_output")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mapping"
    worksheet.append(fieldnames)
    for row in mapping_data:
        worksheet.append([row[field] for field in fieldnames])

    # Styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    # Formater les en-têtes
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Surligner les lignes attendues pour InDesign
    for row in worksheet.iter_rows(min_row=2):
        if row[2].value == "Oui":  # Colonne "Champ attendu"
            for cell in row:
                cell.fill = green_fill

    # Ajuster la largeur des colonnes
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value_length = len(str(cell.value or ""))
            if value_length > max_length:
                max_length = value_length
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    workbook.save(excel_output)
    workbook.close()

    print(f"Fichier Excel créé : {excel_output}")

    # Afficher un résumé
    print("\nRésumé du mapping :")
    print(f"  - Total de colonnes : {len(mapping_data)}")
    print(
        f"  - Colonnes attendues : {len([m for m in mapping_data if m['Champ attendu'] == 'Oui'])}"
    )
    print(
        f"  - Colonnes non mappées : {len([m for m in mapping_data if m['Colonne CSV Final'] == 'Non mappé'])}"
    )

    # Afficher les colonnes attendues
    print("\nColonnes attendues :")
    for m in mapping_data:
        if m["Champ attendu"] == "Oui":
            print(f"  - {m['Colonne Excel Original']} → {m['Colonne CSV Final']}")


if __name__ == "__main__":
    create_mapping()
