#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Lecture du mapping Excel ID dossier -> image source -> image finale."""

from __future__ import annotations

import re
from collections.abc import Callable, Iterable
from typing import Protocol

import openpyxl

from sircom2026_rules import (
    config_list,
    imageid_column_name,
    is_dossier_id_header,
    is_empty_value,
)


class ImageMappingLogger(Protocol):
    def info(self, message: str) -> None: ...

    def warning(self, message: str) -> None: ...

    def error(self, message: str) -> None: ...


ImageMapping = dict[str, dict[str, str]]


def normalize_column_name(column_name: object) -> str:
    """Normalise un nom de colonne pour identifier les colonnes images."""
    if column_name is None:
        return ""
    return re.sub(r"[^\w]", "", str(column_name).strip().lower())


def find_column_by_rule(
    columns: Iterable[object],
    predicate: Callable[[object], bool],
) -> object | None:
    for column in columns:
        if predicate(column):
            return column
    return None


def find_source_image_column(columns: Iterable[object]) -> object | None:
    """Trouve la colonne qui contient le nom source réellement uploadé."""
    normalized_columns = {
        normalize_column_name(column): column
        for column in columns
        if column is not None
    }
    for candidate in config_list("source_image_column_candidates"):
        normalized_candidate = normalize_column_name(candidate)
        if normalized_candidate in normalized_columns:
            return normalized_columns[normalized_candidate]
    for normalized, original in normalized_columns.items():
        if "photo" in normalized:
            return original
    return None


def is_valid_image_value(value: object) -> bool:
    """Indique si une cellule d'image contient une valeur exploitable."""
    return not is_empty_value(value)


def read_excel_mapping(excel_file: str, logger: ImageMappingLogger) -> ImageMapping:
    """
    Lit le fichier Excel enrichi et crée le mapping ID -> image.

    Le fichier attendu contient une colonne `Dossier ID` normalisée, une colonne
    `imageid`, et éventuellement une colonne photo source. En l'absence de
    colonne photo source, le jeu 2026 utilise `imageid` comme nom source.
    """
    logger.info(f"Lecture du fichier Excel : {excel_file}")

    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        worksheet = workbook.active
        headers = [
            cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        header_indexes = {
            header: index for index, header in enumerate(headers) if header is not None
        }

        id_column = find_column_by_rule(headers, is_dossier_id_header)
        imageid_column = find_column_by_rule(
            headers,
            lambda column: (
                normalize_column_name(column)
                == normalize_column_name(imageid_column_name())
            ),
        )
        if id_column is None or imageid_column is None:
            logger.error("Colonnes id_dossier ou imageid non trouvées")
            logger.error(f"   Colonnes disponibles : {headers}")
            return {}

        source_image_column = find_source_image_column(headers)
        if source_image_column is None:
            logger.warning(
                "Colonne photo source non trouvée : fallback sur imageid pour le jeu 2026"
            )
            source_image_column = imageid_column

        logger.info(f"Colonne ID : {id_column}")
        logger.info(f"Colonne Image source : {source_image_column}")
        logger.info(f"Colonne Image finale : {imageid_column_name()}")
        id_index = header_indexes[id_column]
        source_image_index = header_indexes[source_image_column]
        final_image_index = header_indexes[imageid_column]

        mapping: ImageMapping = {}
        skipped = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            dossier_id = row[id_index]
            source_image_name = row[source_image_index]
            final_image_name = row[final_image_index]

            if (
                dossier_id is not None
                and is_valid_image_value(source_image_name)
                and is_valid_image_value(final_image_name)
            ):
                mapping[str(dossier_id)] = {
                    "source_name": str(source_image_name).strip(),
                    "final_name": str(final_image_name).strip(),
                }

                if len(mapping) <= 5:
                    logger.info(
                        f"  Mapping : {dossier_id} -> {source_image_name} "
                        f"-> {final_image_name}"
                    )
            else:
                skipped += 1

        logger.info(f"Mapping créé avec {len(mapping)} correspondances")
        if skipped > 0:
            logger.info(f"{skipped} lignes ignorées (pas d'image ou image invalide)")

        return mapping

    except Exception as exc:
        logger.error(f"Erreur lors de la lecture du fichier Excel : {exc}")
        return {}
    finally:
        if "workbook" in locals():
            workbook.close()
