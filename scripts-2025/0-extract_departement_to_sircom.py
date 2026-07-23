#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script d'extraction de l'onglet Départements vers Sircom.xlsx
=============================================================

Ce script extrait l'onglet "Départements" du fichier Excel multi-onglets
et le sauvegarde comme fichier Sircom.xlsx pour traitement ultérieur
avec la chaîne de scripts-traitement-sircom.

Auteur : Claude Assistant
Date : 17 septembre 2025
"""

import sys
import argparse
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_SOURCE_FILE = "2025.09.17-Dossier JURY national.xlsx"
DEFAULT_OUTPUT_FILE = "Sircom.xlsx"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extraire l'onglet Départements vers Sircom.xlsx.",
    )
    parser.add_argument(
        "--source",
        default=DEFAULT_SOURCE_FILE,
        help=f"Fichier Excel source (défaut : {DEFAULT_SOURCE_FILE}).",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Fichier Excel de sortie (défaut : {DEFAULT_OUTPUT_FILE}).",
    )
    return parser.parse_args()


def extract_departements_to_sircom(
    source_file=DEFAULT_SOURCE_FILE, output_file=DEFAULT_OUTPUT_FILE
):
    """
    Extrait l'onglet Départements et crée le fichier Sircom.xlsx
    """

    source_path = Path(source_file)
    output_path = Path(output_file)

    print("=" * 60)
    print("EXTRACTION DE L'ONGLET DÉPARTEMENTS VERS SIRCOM.XLSX")
    print("=" * 60)
    print()

    # Vérifier que le fichier source existe
    if not source_path.exists():
        print("ERREUR : Le fichier source n'existe pas :")
        print(f"   {source_path}")
        sys.exit(1)

    print(f"Fichier source : {source_path.name}")

    try:
        # Lire l'onglet Départements
        print("\nLecture de l'onglet 'Départements'...")
        source_workbook = load_workbook(source_path, data_only=True)
        if "Départements" not in source_workbook.sheetnames:
            raise ValueError("Onglet 'Départements' introuvable.")
        source_sheet = source_workbook["Départements"]
        rows = [
            [cell_display_value(cell) for cell in row]
            for row in source_sheet.iter_rows()
        ]
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # Afficher les informations sur les données
        print("\nOnglet lu avec succès :")
        print(f"   - Nombre de lignes : {len(data_rows)}")
        print(f"   - Nombre de colonnes : {len(headers)}")

        # Afficher un aperçu des colonnes
        print("\nAperçu des colonnes :")
        for i, col in enumerate(headers[:5]):
            print(f"   - Colonne {i}: {col}")
        if len(headers) > 5:
            print(f"   ... ({len(headers) - 5} colonnes supplémentaires)")

        # Sauvegarder dans le nouveau fichier
        print(f"\nSauvegarde vers : {output_path.name}")

        # Si le fichier existe déjà, le sauvegarder
        if output_path.exists():
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_path = output_path.with_name(
                f"{output_path.stem}_backup_{timestamp}{output_path.suffix}"
            )
            print(
                f"Le fichier existe déjà, création d'une sauvegarde : {backup_path.name}"
            )
            output_path.rename(backup_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Départements"
        for row in rows:
            output_sheet.append(row)
        output_workbook.save(output_path)
        output_workbook.close()
        source_workbook.close()

        print("\nExtraction terminée avec succès !")
        print("   Le fichier Sircom.xlsx est prêt pour le traitement.")

        # Vérification finale
        verif_workbook = load_workbook(output_path, data_only=True)
        verif_sheet = verif_workbook["Départements"]
        print("\nVérification du fichier créé :")
        print(f"   - Lignes : {max(verif_sheet.max_row - 1, 0)}")
        print(f"   - Colonnes : {verif_sheet.max_column}")

        if max(verif_sheet.max_row - 1, 0) == len(
            data_rows
        ) and verif_sheet.max_column == len(headers):
            print("   Intégrité des données confirmée")
        else:
            print("   Attention : différence détectée dans les dimensions")
        verif_workbook.close()

    except Exception as e:
        print(f"\nERREUR lors de l'extraction : {e}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("Prochaine étape :")
    print("   Copier Sircom.xlsx dans scripts-traitement-sircom/")
    print("   puis lancer : python3 sircom_master_script.py --verbose")
    print("=" * 60)


def cell_display_value(cell):
    value = cell.value
    if value is None:
        return None
    if isinstance(value, (int, float)) and float(value).is_integer():
        number_format = str(cell.number_format or "")
        if re.fullmatch(r"0+", number_format):
            return str(int(value)).zfill(len(number_format))
    return value


if __name__ == "__main__":
    args = parse_args()
    extract_departements_to_sircom(args.source, args.output)
