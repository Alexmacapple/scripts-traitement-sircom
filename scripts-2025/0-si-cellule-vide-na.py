#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "Sircom.xlsx", de remplir les cellules vides avec "#N/A",
# de conserver le formatage initial et d'enregistrer le fichier avec le suffixe "_vide_na"

# Créer un environnement virtuel
# python3 -m venv venv

# Activer l'environnement virtuel
# source venv/bin/activate

# Installer les modules
# pip3 install openpyxl

# Lancer le script
# python 0-si-cellule-vide-na.py


import openpyxl
import os

# 1. Définir le chemin du fichier (dans le répertoire courant)
file_path = "Sircom.xlsx"

# Vérifier que le fichier existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    exit(1)

print(f"Traitement du fichier : {file_path}")

# 2. Charger le fichier original avec openpyxl pour conserver formatage et valeurs texte
input_workbook = openpyxl.load_workbook(file_path)

# 3. Créer le nom de fichier de sortie avec le suffixe '_vide_na'
filename, extension = os.path.splitext(file_path)
output_filename = f"{filename}_vide_na{extension}"

# 4. Remplir les cellules vides sans convertir les identifiants en nombres
for sheet_name in input_workbook.sheetnames:
    input_sheet = input_workbook[sheet_name]

    print(f"Mise à jour de la feuille : {sheet_name}")

    for row in input_sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                cell.value = "#N/A"

# 5. Sauvegarder le nouveau fichier Excel avec les données modifiées et le formatage original
input_workbook.save(output_filename)
input_workbook.close()

print(f"Fichier sauvegardé sous : {output_filename}")
print("Traitement terminé avec succès !")
