#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script GÉNÉRIQUE de traitement d'images pour Made in France 2025

VERSION CORRIGÉE - 20 août 2025
================================

HISTORIQUE DU BUG CORRIGÉ :
- La version originale (avant le 20/08/2025) faisait une association par POSITION
- Elle prenait les images dans l'ordre alphabétique et les associait aux lignes du CSV dans l'ordre
- Résultat : mauvaises images sur les mauvais dossiers (ex: bougie sur fiche pizza)

CORRECTION APPORTÉE :
- Le script lit maintenant DIRECTEMENT le fichier Excel source (Sircom.xlsx)
- Il utilise la VRAIE correspondance entre ID et nom d'image source
- Plus d'association aveugle par position !

CE QUE FAIT CE SCRIPT :
1. Lit le fichier Excel source (Sircom.xlsx)
2. Récupère le mapping réel :
   - Colonne f_id : ID du dossier (ex: 24331205)
   - Colonne photo source : nom réel de l'image uploadée (ex: packshot.jpg)
   - Colonne imageid : nom final attendu (ex: dossier-24331205.jpg)
3. Pour chaque dossier, cherche L'IMAGE EXACTE par son nom
4. La renomme selon le pattern : dossier-{ID}.jpg
5. Redimensionne et optimise pour l'impression (350px, 300 DPI)

GÉNÉRICITÉ :
- Fonctionne avec n'importe quel nouveau lot d'images
- Gère les cas spéciaux (extensions différentes, noms approchants)
- Signale les images manquantes ou non référencées
"""

import os
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime
from PIL import Image
import openpyxl
import re

# ==========================================
# CONFIGURATION
# ==========================================

# Paramètres de traitement d'images
MAX_WIDTH = 350
JPEG_QUALITY = 100
DPI = 300

# Chemins des fichiers et répertoires
EXCEL_FILE = "7-add-pathimg.xlsx"  # Utiliser le fichier avec le mapping
SOURCE_DIR = os.environ.get("SIRCOM_SOURCE_IMAGE_DIR", "images")
TARGET_DIR = "export_images_id_dossier_rename_resize"

# Extensions d'images supportées
ALLOWED_EXTENSIONS = [
    'jpg', 'jpeg', 'png', 'gif', 'webp', 'tif', 'tiff', 
    'bmp', 'eps', 'svg', 'ico', 'heic', 'heif', 'psd', 
    'raw', 'hdr', 'exr', 'jp2', 'pgm', 'ppm', 'xcf'
]

SOURCE_IMAGE_COLUMN_CANDIDATES = [
    "y_photodu",
    "y_photo",
    "photo_du_produit",
    "photo_produit",
    "photo",
    "nom_image_source",
    "image_source",
    "source_image",
]

# ==========================================
# CONFIGURATION DU LOGGING
# ==========================================

def setup_logging():
    """Configure le système de logs pour le traitement d'images"""
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    log_filename = f"images-processing-{timestamp}.log"
    
    # Configuration du logger
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("="*80)
    logger.info("TRAITEMENT GÉNÉRIQUE DES IMAGES - MADE IN FRANCE 2025")
    logger.info("="*80)
    logger.info(f"Date/heure de début : {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}")
    logger.info(f"Fichier de log : {log_filename}")
    
    return logger

# ==========================================
# FONCTIONS UTILITAIRES
# ==========================================

def normalize_filename(filename):
    """Normalise un nom de fichier pour la comparaison"""
    if not filename:
        return ""
    # Enlever les espaces multiples et normaliser
    return re.sub(r'\s+', ' ', filename.strip())

def normalize_column_name(column_name):
    """Normalise un nom de colonne pour identifier la colonne photo source"""
    if column_name is None:
        return ""
    return re.sub(r"[^\w]", "", str(column_name).strip().lower())

def find_source_image_column(columns):
    """Trouve la colonne qui contient le nom source réellement uploadé"""
    normalized_columns = {
        normalize_column_name(column): column
        for column in columns
    }
    for candidate in SOURCE_IMAGE_COLUMN_CANDIDATES:
        normalized_candidate = normalize_column_name(candidate)
        if normalized_candidate in normalized_columns:
            return normalized_columns[normalized_candidate]
    for normalized, original in normalized_columns.items():
        if "photo" in normalized:
            return original
    return None

def is_valid_image_value(value):
    """Indique si une cellule d'image contient une valeur exploitable"""
    if value is None:
        return False
    text = str(value).strip()
    return text not in ["", "#N/A", "N/A"]

class AmbiguousImageMatchError(ValueError):
    """Signalée quand plusieurs fichiers peuvent correspondre à la même image."""

    def __init__(self, target_name, candidates):
        self.target_name = target_name
        self.candidates = sorted(candidates)
        super().__init__(
            f"Plusieurs images correspondent à {target_name!r} : "
            + ", ".join(self.candidates)
        )


def find_best_match(target_name, available_files, logger):
    """
    Trouve la meilleure correspondance pour un nom de fichier
    Gère les cas où les noms ne correspondent pas exactement
    """
    target_normalized = normalize_filename(target_name)
    
    # 1. Correspondance exacte
    exact_candidates = [filename for filename in available_files if filename == target_name]
    if len(exact_candidates) == 1:
        return target_name
    if len(exact_candidates) > 1:
        raise AmbiguousImageMatchError(target_name, exact_candidates)
    
    # 2. Correspondance normalisée
    normalized_candidates = [
        filename for filename in available_files if normalize_filename(filename) == target_normalized
    ]
    if len(normalized_candidates) == 1:
        logger.info(f"  Correspondance normalisée trouvée : {normalized_candidates[0]}")
        return normalized_candidates[0]
    if len(normalized_candidates) > 1:
        raise AmbiguousImageMatchError(target_name, normalized_candidates)
    
    # 3. Correspondance sans extension (pour les erreurs .pptx → .jpg)
    target_base = os.path.splitext(target_normalized)[0]
    base_candidates = [
        filename
        for filename in available_files
        if os.path.splitext(normalize_filename(filename))[0] == target_base
    ]
    if len(base_candidates) == 1:
        logger.info(f"  Correspondance par nom de base trouvée : {base_candidates[0]}")
        return base_candidates[0]
    if len(base_candidates) > 1:
        raise AmbiguousImageMatchError(target_name, base_candidates)
    
    # 4. Correspondance partielle (contient le nom principal)
    # Utile pour les cas comme "Virebent_photophore" vs "MadeByVirebent_2025_..."
    if len(target_base) > 10:  # Seulement pour les noms assez longs
        keywords = target_base.split('_')[:2]  # Prendre les 2 premiers mots
        if keywords:
            main_keyword = keywords[0]
            candidates = [
                filename
                for filename in available_files
                if main_keyword.lower() in filename.lower()
            ]
            if len(candidates) == 1:
                logger.info(f"  Correspondance partielle trouvée : {candidates[0]}")
                return candidates[0]
            if len(candidates) > 1:
                raise AmbiguousImageMatchError(target_name, candidates)
    
    return None

def check_prerequisites(logger):
    """Vérifier les prérequis"""
    logger.info("VÉRIFICATION DES PRÉREQUIS")
    
    # Vérifier le fichier Excel
    if not os.path.exists(EXCEL_FILE):
        logger.error(f"Fichier Excel manquant : {EXCEL_FILE}")
        return False
    logger.info(f"Fichier Excel trouvé : {EXCEL_FILE}")
    
    # Vérifier le répertoire source des images
    if not os.path.exists(SOURCE_DIR):
        logger.error(f"Répertoire d'images manquant : {SOURCE_DIR}")
        logger.error("Veuillez créer le répertoire et y placer les images à traiter")
        return False
    logger.info(f"Répertoire d'images trouvé : {SOURCE_DIR}")
    
    # Vérifier qu'il y a des images dans le répertoire
    image_files = get_available_images(SOURCE_DIR)
    if not image_files:
        logger.error(f"Aucune image trouvée dans : {SOURCE_DIR}")
        logger.error(f"Extensions supportées : {', '.join(ALLOWED_EXTENSIONS)}")
        return False
    logger.info(f"{len(image_files)} images trouvées dans le répertoire source")
    
    return True

def read_excel_mapping(excel_file, logger):
    """
    Lit le fichier Excel et crée le mapping ID -> nom d'image

    Adapté pour lire depuis 7-add-pathimg.xlsx qui contient :
    - Colonne f_id : ID du dossier
    - Colonne photo source : nom réel de l'image uploadée
    - Colonne imageid : nom final normalisé

    Retourne un dictionnaire :
    {ID_dossier: {"source_name": nom_image_source, "final_name": nom_image_final}}
    """
    logger.info(f"Lecture du fichier Excel : {excel_file}")

    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        worksheet = workbook.active
        headers = [
            cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        header_indexes = {
            header: index
            for index, header in enumerate(headers)
            if header is not None
        }

        # Utiliser les colonnes par nom (plus robuste)
        if 'f_id' not in header_indexes or 'imageid' not in header_indexes:
            logger.error("Colonnes f_id ou imageid non trouvées")
            logger.error(f"   Colonnes disponibles : {headers}")
            return {}
        source_image_column = find_source_image_column(headers)
        if source_image_column is None:
            logger.error("Colonne photo source non trouvée")
            logger.error(f"   Colonnes disponibles : {headers}")
            return {}

        logger.info("Colonne ID : f_id")
        logger.info(f"Colonne Image source : {source_image_column}")
        logger.info("Colonne Image finale : imageid")
        id_index = header_indexes['f_id']
        source_image_index = header_indexes[source_image_column]
        final_image_index = header_indexes['imageid']

        # Créer le mapping
        mapping = {}
        skipped = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            dossier_id = row[id_index]
            source_image_name = row[source_image_index]
            final_image_name = row[final_image_index]

            # Vérifier que les deux valeurs sont valides
            if (
                dossier_id is not None
                and is_valid_image_value(source_image_name)
                and is_valid_image_value(final_image_name)
            ):
                mapping[str(dossier_id)] = {
                    "source_name": str(source_image_name).strip(),
                    "final_name": str(final_image_name).strip(),
                }

                if len(mapping) <= 5:  # Afficher les 5 premiers
                    logger.info(
                        f"  Mapping : {dossier_id} → {source_image_name} "
                        f"→ {final_image_name}"
                    )
            else:
                skipped += 1

        logger.info(f"Mapping créé avec {len(mapping)} correspondances")
        if skipped > 0:
            logger.info(f"{skipped} lignes ignorées (pas d'image ou image invalide)")

        return mapping

    except Exception as e:
        logger.error(f"Erreur lors de la lecture du fichier Excel : {e}")
        return {}
    finally:
        if 'workbook' in locals():
            workbook.close()

def get_available_images(source_dir):
    """Récupère la liste des images disponibles"""
    image_files = {}
    
    for f in os.listdir(source_dir):
        # Ignorer les fichiers cachés
        if f.startswith('.'):
            continue
            
        ext = os.path.splitext(f)[1][1:].lower()
        if ext in ALLOWED_EXTENSIONS:
            image_files[f] = os.path.join(source_dir, f)
    
    return image_files

def process_and_rename_image(image_path, new_name, target_dir, logger):
    """Traite une image (redimensionnement, conversion, DPI) et la renomme"""
    try:
        # Augmente la limite de pixels maximum pour l'image PIL
        Image.MAX_IMAGE_PIXELS = None
        
        # Ouvre l'image
        with Image.open(image_path) as im:
            # Redimensionne l'image en conservant l'homotétie
            im.thumbnail((MAX_WIDTH, MAX_WIDTH), Image.Resampling.LANCZOS)
            
            # Ajoute un fond blanc à une image avec un fond transparent
            if im.mode == 'RGBA':
                background = Image.new('RGB', im.size, (255, 255, 255))
                background.paste(im, mask=im.split()[3])
                im = background
            
            # Convertit en RGB si nécessaire
            if im.mode != 'RGB':
                im = im.convert('RGB')
            
            # Enregistre l'image traitée avec le nouveau nom
            out_path = os.path.join(target_dir, new_name)
            im.save(out_path, 'JPEG', quality=JPEG_QUALITY, dpi=(DPI, DPI))
            
            # Vérifier que le fichier a été créé
            if os.path.exists(out_path):
                file_size = os.path.getsize(out_path)
                logger.info(f"  Image traitée : {new_name} ({file_size:,} octets)")
                return True, file_size
            else:
                logger.error(f"  Fichier de sortie non créé : {new_name}")
                return False, 0
            
    except Exception as e:
        logger.error(f"  ERREUR lors du traitement : {e}")
        return False, 0

def parse_args():
    parser = argparse.ArgumentParser(
        description="Traiter les images source Made in France 2025."
    )
    parser.add_argument(
        "--source-dir",
        default=os.environ.get("SIRCOM_SOURCE_IMAGE_DIR", SOURCE_DIR),
        help="Dossier contenant les images source",
    )
    parser.add_argument(
        "--target-dir",
        default=TARGET_DIR,
        help="Dossier de sortie des images JPG traitées",
    )
    parser.add_argument(
        "--excel-file",
        default=EXCEL_FILE,
        help="Fichier Excel de mapping image à lire",
    )
    return parser.parse_args()


def main():
    """Fonction principale"""
    global SOURCE_DIR, TARGET_DIR, EXCEL_FILE

    args = parse_args()
    source_dir = args.source_dir
    target_dir = args.target_dir
    excel_file = args.excel_file
    SOURCE_DIR = source_dir
    TARGET_DIR = target_dir
    EXCEL_FILE = excel_file
    
    # Configuration du logging
    logger = setup_logging()
    
    try:
        # En-tête
        logger.info("PARAMÈTRES DE TRAITEMENT :")
        logger.info(f"  - Largeur max : {MAX_WIDTH}px")
        logger.info(f"  - Qualité JPEG : {JPEG_QUALITY}%")
        logger.info(f"  - DPI : {DPI}")
        logger.info(f"  - Source : {source_dir}")
        logger.info(f"  - Destination : {target_dir}")
        
        # Vérifications préliminaires
        if not check_prerequisites(logger):
            logger.error("Prérequis non satisfaits - Arrêt du traitement")
            return False
        
        # Lecture du mapping depuis Excel
        mapping = read_excel_mapping(excel_file, logger)
        if not mapping:
            logger.error("Impossible de continuer sans mapping ID/image")
            return False
        
        # Récupération des images disponibles
        available_images = get_available_images(source_dir)
        logger.info(f"{len(available_images)} images disponibles dans {source_dir}")
        
        # Créer le répertoire de destination
        Path(target_dir).mkdir(parents=True, exist_ok=True)
        logger.info(f"Répertoire de destination : {target_dir}")
        
        # Traitement des images selon le mapping
        logger.info("="*80)
        logger.info("DÉBUT DU TRAITEMENT ET RENOMMAGE DES IMAGES")
        logger.info("="*80)
        
        success_count = 0
        error_count = 0
        not_found_count = 0
        total_size = 0
        
        for dossier_id, image_mapping in mapping.items():
            source_image_name = image_mapping["source_name"]
            new_name = image_mapping["final_name"]

            # CORRECTION IMPORTANTE : On cherche l'image PAR SON NOM, pas par position !
            # L'ancien script prenait juste l'image N pour le dossier N (FAUX)
            # Maintenant on cherche l'image qui a VRAIMENT ce nom
            try:
                matched_file = find_best_match(source_image_name, available_images.keys(), logger)
            except AmbiguousImageMatchError as exc:
                logger.error(
                    "Image ambiguë : %s (dossier %s). Candidats : %s",
                    source_image_name,
                    dossier_id,
                    ", ".join(exc.candidates),
                )
                error_count += 1
                continue
            
            if matched_file:
                logger.info(f"Traitement : {matched_file} → {new_name}")
                logger.info(f"  ID Dossier : {dossier_id}")
                if matched_file != source_image_name:
                    logger.info(f"  Nom dans Excel : {source_image_name}")
                
                source_path = available_images[matched_file]
                
                # Traiter et renommer l'image
                success, file_size = process_and_rename_image(source_path, new_name, target_dir, logger)
                if success:
                    success_count += 1
                    total_size += file_size
                else:
                    error_count += 1
            else:
                logger.warning(f"Image non trouvée : {source_image_name} (dossier {dossier_id})")
                not_found_count += 1
        
        # Images non utilisées
        used_images = set()
        for image_mapping in mapping.values():
            source_name = image_mapping["source_name"]
            try:
                matched = find_best_match(source_name, available_images.keys(), logger)
            except AmbiguousImageMatchError:
                matched = None
            if matched:
                used_images.add(matched)
        
        unused_images = [img for img in available_images.keys() if img not in used_images]
        if unused_images:
            logger.warning("IMAGES NON RÉFÉRENCÉES DANS EXCEL (non traitées) :")
            for img in unused_images:
                logger.warning(f"  {img}")
        
        # Résumé final
        logger.info("="*80)
        logger.info("RÉSUMÉ DU TRAITEMENT")
        logger.info("="*80)
        logger.info(f"Images traitées avec succès : {success_count}")
        logger.info(f"Erreurs de traitement : {error_count}")
        logger.info(f"Images référencées mais non trouvées : {not_found_count}")
        logger.info(f"Répertoire de sortie : {target_dir}")
        logger.info(f"Espace disque utilisé : {total_size:,} octets ({total_size/1024/1024:.2f} MB)")
        
        # Lister les fichiers dans le répertoire de destination
        if os.path.exists(target_dir):
            output_files = sorted(os.listdir(target_dir))
            logger.info(f"\nFICHIERS CRÉÉS DANS {target_dir} :")
            for file in output_files[:10]:  # Afficher les 10 premiers
                logger.info(f"  - {file}")
            if len(output_files) > 10:
                logger.info(f"  ... et {len(output_files) - 10} autres fichiers")
            logger.info(f"Total fichiers créés : {len(output_files)}")
        
        # Déterminer le succès
        if success_count > 0 and error_count == 0 and not_found_count == 0:
            logger.info("TRAITEMENT TERMINÉ AVEC SUCCÈS !")
            return True
        if success_count > 0:
            logger.error("TRAITEMENT INCOMPLET : erreurs ou images manquantes")
        else:
            logger.error("AUCUNE IMAGE N'A PU ÊTRE TRAITÉE")
        return False
            
    except Exception as e:
        logger.error(f"Erreur critique : {e}")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
