#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "1-header-lettres-colonne-excel-mapping-excel.xlsx",
# d'ajouter une colonne "image-id" juste après la colonne "B_ID",
# de générer des noms de fichiers basés sur les valeurs de la colonne "B_ID" (format: "dossier-[ID].jpg"),
# et d'enregistrer le fichier sous "3-image-id-adder-excel-fusion.xlsx"

# python3 2-image_id_adder.py

import openpyxl
import os

# 1. Définir le fichier source
file_path = "1-header-lettres-colonne-excel-mapping-excel.xlsx"

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print("Assurez-vous d'avoir exécuté le script '1-header_lettres_colonne.py' au préalable.")
    exit(1)

print(f"Traitement du fichier : {file_path}")

# 3. Ouvrir le fichier Excel
try:
    workbook = openpyxl.load_workbook(file_path)
    print(f"Fichier ouvert avec succès")
except Exception as e:
    print(f"Erreur lors de l'ouverture du fichier : {e}")
    exit(1)

try:
    # 4. Sélectionner la première feuille (peu importe son nom)
    worksheet = workbook.active
    sheet_name = worksheet.title
    print(f"Traitement de la feuille : {sheet_name}")

    # 5. Chercher la colonne "F_ID"
    id_column = None
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == "F_ID":
                id_column = cell.column
                print(f"Colonne 'F_ID' trouvée en position : {openpyxl.utils.get_column_letter(id_column)}")
                break
        if id_column:
            break

    if id_column:
        # 6. Insérer une nouvelle colonne "image-id" juste après "F_ID"
        image_id_column = id_column + 1
        worksheet.insert_cols(image_id_column)
        worksheet.cell(row=1, column=image_id_column).value = "image-id"
        print(f"Colonne 'image-id' ajoutée en position : {openpyxl.utils.get_column_letter(image_id_column)}")

        # 7. Parcourir les données de la colonne "F_ID" et générer le contenu de la colonne "image-id"
        rows_processed = 0
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            cell = row[id_column - 1]  # Colonne F_ID
            if cell.value is not None and cell.value != "#N/A":
                file_name = f"dossier-{cell.value}.jpg"
                worksheet.cell(row=cell.row, column=image_id_column).value = file_name
                rows_processed += 1
                if rows_processed <= 5:  # Afficher les 5 premiers exemples
                    print(f"  Ligne {cell.row}: ID={cell.value} → {file_name}")
        
        print(f"{rows_processed} noms de fichiers générés")

        # 8. Enregistrer les modifications dans le fichier Excel
        output_filename = "2-image-id-adder-excel-fusion.xlsx"
        workbook.save(output_filename)
        print(f"Fichier sauvegardé sous : {output_filename}")
        print("Modification terminée avec succès !")
        
    else:
        print("Erreur : La colonne 'F_ID' n'a pas été trouvée dans la feuille.")
        print("Vérifiez que le script précédent a bien été exécuté.")
        exit(1)

except Exception as e:
    print(f"Erreur lors du traitement : {e}")
    exit(1)
finally:
    # 9. Fermer le fichier Excel
    workbook.close()
    print("Fichier fermé")