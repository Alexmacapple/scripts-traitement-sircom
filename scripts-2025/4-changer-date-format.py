#!/usr/bin/env python3

### Ce script permet :

# Formate les colonnes de date en texte JJ/MM/AAAA quand elles existent.

import os
import re
import unicodedata
from datetime import date, datetime

import openpyxl

# 1. Définir les fichiers source et destination
input_file = "3-fusion-tri-region-departement.xlsx"
output_file = "4-changer-date.xlsx"


def normalize_header(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("utf-8")
    return re.sub(r"[^a-z0-9]", "", text.lower())


def date_column_indexes(sheet):
    indexes = []
    for cell in sheet[1]:
        normalized = normalize_header(cell.value)
        if "date" in normalized:
            indexes.append(cell.column)
    return indexes


def format_date_value(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "#N/A":
            return value
        for pattern in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, pattern).strftime("%d/%m/%Y")
            except ValueError:
                continue
    return value


def format_date_columns(sheet):
    indexes = date_column_indexes(sheet)
    changed_count = 0
    for column_index in indexes:
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            formatted_value = format_date_value(cell.value)
            if formatted_value != cell.value:
                cell.value = formatted_value
                changed_count += 1
    return indexes, changed_count


# 2. Vérifier que le fichier source existe
if not os.path.exists(input_file):
    print(
        f"Erreur : Le fichier '{input_file}' n'existe pas dans le répertoire courant."
    )
    print(
        "Assurez-vous d'avoir exécuté le script '3-fusion_tri_region_departement.py' au préalable."
    )
    exit(1)

print(f"Traitement du fichier : {input_file}")

try:
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    indexes, changed_count = format_date_columns(sheet)
    workbook.save(output_file)
    workbook.close()
    print(f"Fichier sauvegardé : {output_file}")
    if indexes:
        print(f"Colonnes de date détectées : {len(indexes)}")
        print(f"Cellules de date formatées : {changed_count}")
    else:
        print("Aucune colonne de date détectée : contenu conservé.")
    print("Opération terminée avec succès !")
except Exception as e:
    print(f"Erreur lors du formatage des dates : {e}")
    exit(1)
