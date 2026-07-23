#!/usr/bin/env python3

### Ce script permet :

# de traiter automatiquement le fichier "4-changer-date.xlsx",
# de créer une copie finale sous le nom "5-livrable-final-word.xlsx",
# de conserver le nom d'onglet original,
# et de préparer le fichier final pour utilisation

# python3 6-livrable-final.py

import openpyxl
import os

# 1. Définir le fichier source
file_path = "4-changer-date.xlsx"

# 2. Vérifier que le fichier source existe
if not os.path.exists(file_path):
    print(f"Erreur : Le fichier '{file_path}' n'existe pas dans le répertoire courant.")
    print("Assurez-vous d'avoir exécuté le script '5-changer-date-format.py' au préalable.")
    exit(1)

print(f"Traitement du fichier source : {file_path}")

try:
    # 3. Charger le fichier Excel source
    workbook = openpyxl.load_workbook(file_path)
    print("Fichier source ouvert avec succès")
    
    # Récupérer des informations sur le fichier source
    worksheet = workbook.active
    original_sheet_name = worksheet.title
    rows_count = worksheet.max_row
    columns_count = worksheet.max_column
    
    print("Informations du fichier source :")
    print(f"  - Nom de l'onglet : '{original_sheet_name}'")
    print(f"  - Nombre de lignes : {rows_count}")
    print(f"  - Nombre de colonnes : {columns_count}")

    # 4. Définir le nom du fichier de sortie
    output_file_name = '5-livrable-final-word.xlsx'
    
    print("\nCréation du livrable final...")
    
    # 5. Créer une copie du fichier avec le nouveau nom
    workbook.save(output_file_name)
    print(f"Copie créée sous le nom : {output_file_name}")
    
    # 6. Charger la copie pour vérification
    workbook_copy = openpyxl.load_workbook(output_file_name)
    worksheet_copy = workbook_copy.active
    
    print("Livrable final vérifié :")
    print(f"  - Nom de l'onglet conservé : '{worksheet_copy.title}'")
    print(f"  - Données préservées : {worksheet_copy.max_row} lignes, {worksheet_copy.max_column} colonnes")
    
    # 7. Enregistrer et fermer les fichiers
    workbook_copy.save(output_file_name)
    workbook_copy.close()
    workbook.close()
    
    print("\nLivrable final créé avec succès !")
    print(f"Fichier disponible : {output_file_name}")
    print("Résumé :")
    print(f"  ├── Fichier source : {file_path}")
    print(f"  ├── Fichier final : {output_file_name}")
    print(f"  ├── Onglet : {original_sheet_name}")
    print(f"  └── Contenu : {rows_count-1} dossiers + en-têtes")

except FileNotFoundError:
    print(f"Erreur : Le fichier '{file_path}' est introuvable.")
    exit(1)
except PermissionError:
    print(f"Erreur : Permission refusée. Vérifiez que le fichier '{output_file_name}' n'est pas ouvert dans Excel.")
    exit(1)
except Exception as e:
    print(f"Erreur lors du traitement : {e}")
    exit(1)

print("\nLe livrable final est prêt pour utilisation !")