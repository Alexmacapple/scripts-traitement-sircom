"""Excel structure diagnostics for the Sircom 2026 import flow.

The diagnostic intentionally reports workbook structure, headers and counts, not
business cell values. It is meant to be reused by the future FastAPI upload
endpoint and by the local CLI.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from sircom2026.config import (
    DEFAULT_MAX_EXCEL_CELLS,
    DEFAULT_MAX_EXCEL_COLUMNS,
    DEFAULT_MAX_EXCEL_ROWS,
)


ID_HEADER_RE = re.compile(r"^(id|id[_\s-]*dossier|dossier[_\s-]*id)$", re.I)
REGION_RE = re.compile(r"\bregion\b|\brégion\b", re.I)
DEPARTMENT_RE = re.compile(r"departement|département|code\s*postal", re.I)
DATE_RE = re.compile(
    r"\bdate\b|\bcréé\b|\bcree\b|\bdéposé\b|\bdepose\b|\btraité\b|\btraite\b|\ble$",
    re.I,
)
IMAGE_RE = re.compile(r"photo|image|logo|piece|pièce|fichier", re.I)
SENSITIVE_RE = re.compile(
    r"id|siret|siren|téléphone|telephone|code\s*postal|département|departement|code\s+insee|rna|tva",
    re.I,
)
EXCEL_DIMENSIONS_EXCEEDED_CODE = "SIRCOM_EXCEL_DIMENSIONS_EXCEEDED"
HEADER_SCAN_ROWS = 100


@dataclass(frozen=True)
class ExcelDimensionLimits:
    max_rows: int
    max_columns: int
    max_cells: int


@dataclass(frozen=True)
class ExcelDimensionLimitViolation:
    limit_exceeded: str
    sheet: str
    observed: int
    maximum: int

    def public_details(self) -> dict[str, Any]:
        return {
            "limit_exceeded": self.limit_exceeded,
            "sheet": self.sheet,
            "observed": self.observed,
            "max": self.maximum,
        }


class ExcelDimensionLimitError(ValueError):
    def __init__(self, violation: ExcelDimensionLimitViolation) -> None:
        super().__init__(
            f"{violation.sheet}: {violation.limit_exceeded} "
            f"{violation.observed} > {violation.maximum}"
        )
        self.violation = violation


DEFAULT_EXCEL_DIMENSION_LIMITS = ExcelDimensionLimits(
    max_rows=DEFAULT_MAX_EXCEL_ROWS,
    max_columns=DEFAULT_MAX_EXCEL_COLUMNS,
    max_cells=DEFAULT_MAX_EXCEL_CELLS,
)


@dataclass
class ColumnCandidate:
    column: str
    header: str
    non_empty_values: int | None = None
    unique_values: int | None = None
    duplicate_values: int | None = None
    blank_values: int | None = None


@dataclass
class SheetDiagnostic:
    name: str
    state: str
    rows: int
    columns: int
    ignored: bool = False
    ignore_reason: str | None = None
    importable: bool = True
    blockers: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    header_row: int | None = None
    non_empty_headers: int = 0
    empty_header_columns_with_data: list[str] = field(default_factory=list)
    duplicate_headers: list[str] = field(default_factory=list)
    cleaned_header_collisions: list[str] = field(default_factory=list)
    id_candidates: list[ColumnCandidate] = field(default_factory=list)
    region_candidates: list[ColumnCandidate] = field(default_factory=list)
    department_candidates: list[ColumnCandidate] = field(default_factory=list)
    date_candidates: list[ColumnCandidate] = field(default_factory=list)
    image_candidates: list[ColumnCandidate] = field(default_factory=list)
    sensitive_candidates: list[ColumnCandidate] = field(default_factory=list)
    source_headers: list[ColumnCandidate] = field(default_factory=list)
    hidden_columns: list[str] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_data_rows: list[int] = field(default_factory=list)
    merged_ranges: list[str] = field(default_factory=list)
    formula_cells_sample: list[str] = field(default_factory=list)
    headers_preview: list[str] = field(default_factory=list)
    dimension_limits_exceeded: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class WorkbookDiagnostic:
    path: str
    filename: str
    sheet_count: int
    importable: bool
    blockers: list[str]
    warnings: list[str]
    sheets: list[SheetDiagnostic]
    cleaned_header_collisions: dict[str, list[str]] = field(default_factory=dict)


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def ascii_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_value.lower().strip()


def clean_indesign_header(value: str) -> str:
    cleaned = ascii_key(value)
    cleaned = re.sub(r"[^\w]", "", cleaned)
    return cleaned[:10]


def is_hidden_row(ws, row: int) -> bool:
    return bool(ws.row_dimensions[row].hidden)


def row_has_non_empty_value(ws, row: int) -> bool:
    return any(
        normalize_header(ws.cell(row=row, column=column).value)
        for column in range(1, ws.max_column + 1)
    )


def visible_non_empty_data_rows(ws, header_row: int) -> Iterable[int]:
    for row in range(header_row + 1, ws.max_row + 1):
        if is_hidden_row(ws, row):
            continue
        if row_has_non_empty_value(ws, row):
            yield row


def hidden_non_empty_data_rows(ws, header_row: int) -> list[int]:
    return [
        row
        for row in range(header_row + 1, ws.max_row + 1)
        if is_hidden_row(ws, row) and row_has_non_empty_value(ws, row)
    ]


def visible_rows_before_header_have_data(ws, header_row: int) -> bool:
    return any(
        not is_hidden_row(ws, row) and row_has_non_empty_value(ws, row)
        for row in range(1, header_row)
    )


def iter_non_empty_values(
    ws,
    column: int,
    *,
    row_numbers: Iterable[int],
) -> Iterable[str]:
    for row in row_numbers:
        value = normalize_header(ws.cell(row=row, column=column).value)
        if value:
            yield value


def count_non_empty_cells(
    ws,
    limits: ExcelDimensionLimits = DEFAULT_EXCEL_DIMENSION_LIMITS,
) -> int:
    count = 0
    visited = 0
    for row in ws.iter_rows():
        for cell in row:
            visited += 1
            if visited > limits.max_cells:
                raise ExcelDimensionLimitError(
                    ExcelDimensionLimitViolation(
                        limit_exceeded="max_cells",
                        sheet=ws.title,
                        observed=visited,
                        maximum=limits.max_cells,
                    )
                )
            if normalize_header(cell.value):
                count += 1
    return count


def excel_dimension_limits_from_settings(settings: object) -> ExcelDimensionLimits:
    return ExcelDimensionLimits(
        max_rows=int(getattr(settings, "max_excel_rows")),
        max_columns=int(getattr(settings, "max_excel_columns")),
        max_cells=int(getattr(settings, "max_excel_cells")),
    )


def check_worksheet_dimensions(
    ws,
    limits: ExcelDimensionLimits = DEFAULT_EXCEL_DIMENSION_LIMITS,
) -> None:
    rows = _dimension_value(ws.max_row)
    columns = _dimension_value(ws.max_column)
    if rows > limits.max_rows:
        raise ExcelDimensionLimitError(
            ExcelDimensionLimitViolation(
                limit_exceeded="max_rows",
                sheet=ws.title,
                observed=rows,
                maximum=limits.max_rows,
            )
        )
    if columns > limits.max_columns:
        raise ExcelDimensionLimitError(
            ExcelDimensionLimitViolation(
                limit_exceeded="max_columns",
                sheet=ws.title,
                observed=columns,
                maximum=limits.max_columns,
            )
        )
    cells = rows * columns
    if cells > limits.max_cells:
        raise ExcelDimensionLimitError(
            ExcelDimensionLimitViolation(
                limit_exceeded="max_cells",
                sheet=ws.title,
                observed=cells,
                maximum=limits.max_cells,
            )
        )


def _dimension_value(value: int | None) -> int:
    if value is None:
        return 0
    return max(0, int(value))


def detect_header_row(ws, max_scan: int = HEADER_SCAN_ROWS) -> int | None:
    best_row = None
    best_score = 0.0
    for row in range(1, min(max_scan, ws.max_row) + 1):
        if is_hidden_row(ws, row):
            continue
        values = [
            normalize_header(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        ]
        present = [value for value in values if value]
        if len(present) < 2:
            continue
        text_ratio = sum(
            any(char.isalpha() for char in value) for value in present
        ) / len(present)
        unique_ratio = len(set(present)) / len(present)
        density = len(present) / max(1, ws.max_column)
        id_header_bonus = (
            0.75
            if any(ID_HEADER_RE.match(ascii_key(value)) for value in present)
            else 0
        )
        score = text_ratio + unique_ratio + density + id_header_bonus
        if score > best_score:
            best_score = score
            best_row = row
    if best_score < 1.8:
        return None
    return best_row


def hidden_column_ranges(ws) -> list[str]:
    hidden: list[str] = []
    for key, dimension in ws.column_dimensions.items():
        if dimension.hidden:
            hidden.append(key)
    return hidden


def hidden_row_numbers(ws) -> list[int]:
    return [index for index, dimension in ws.row_dimensions.items() if dimension.hidden]


def formula_cell_sample(ws, limit: int = 20) -> list[str]:
    sample: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                sample.append(cell.coordinate)
                if len(sample) >= limit:
                    return sample
    return sample


def make_id_candidate(ws, column: int, header: str, header_row: int) -> ColumnCandidate:
    importable_rows = list(visible_non_empty_data_rows(ws, header_row))
    values = list(iter_non_empty_values(ws, column, row_numbers=importable_rows))
    unique = set(values)
    return ColumnCandidate(
        column=get_column_letter(column),
        header=header,
        non_empty_values=len(values),
        unique_values=len(unique),
        duplicate_values=len(values) - len(unique),
        blank_values=len(importable_rows) - len(values),
    )


def make_simple_candidate(column: int, header: str) -> ColumnCandidate:
    return ColumnCandidate(column=get_column_letter(column), header=header)


def diagnose_sheet(
    ws,
    limits: ExcelDimensionLimits = DEFAULT_EXCEL_DIMENSION_LIMITS,
) -> SheetDiagnostic:
    diagnostic = SheetDiagnostic(
        name=ws.title,
        state=ws.sheet_state,
        rows=ws.max_row,
        columns=ws.max_column,
    )

    try:
        check_worksheet_dimensions(ws, limits)
        non_empty_cells = count_non_empty_cells(ws, limits)
    except ExcelDimensionLimitError as exc:
        _block_sheet_for_dimension_limit(diagnostic, exc.violation)
        return diagnostic

    if non_empty_cells == 0:
        diagnostic.ignored = True
        diagnostic.ignore_reason = "empty sheet"
        diagnostic.importable = True
        diagnostic.warnings.append("Onglet vide ignoré.")
        return diagnostic

    diagnostic.hidden_columns = hidden_column_ranges(ws)
    diagnostic.hidden_rows = hidden_row_numbers(ws)
    diagnostic.merged_ranges = [
        str(cell_range) for cell_range in ws.merged_cells.ranges
    ]
    diagnostic.formula_cells_sample = formula_cell_sample(ws)

    if ws.sheet_state != "visible":
        diagnostic.blockers.append("Onglet masqué.")
    if diagnostic.hidden_columns:
        diagnostic.blockers.append(
            f"{len(diagnostic.hidden_columns)} colonne(s) masquée(s)."
        )
    if diagnostic.merged_ranges:
        diagnostic.blockers.append(
            f"{len(diagnostic.merged_ranges)} cellule(s) fusionnée(s)."
        )
    if diagnostic.formula_cells_sample:
        diagnostic.blockers.append("Formules détectées.")

    header_row = detect_header_row(ws)
    diagnostic.header_row = header_row
    if header_row is None:
        diagnostic.blockers.append("En-tête non détecté.")
        diagnostic.blockers.append("Clé primaire dossier non détectée.")
        diagnostic.importable = False
        return diagnostic
    if header_row != 1:
        if visible_rows_before_header_have_data(ws, header_row):
            diagnostic.blockers.append("En-tête détecté hors première ligne.")
        else:
            diagnostic.warnings.append("Ligne(s) vide(s) avant l'en-tête ignorée(s).")
    diagnostic.hidden_data_rows = hidden_non_empty_data_rows(ws, header_row)
    if diagnostic.hidden_data_rows:
        diagnostic.warnings.append("Ligne(s) masquée(s) ignorée(s) à l'import.")

    headers = [
        normalize_header(ws.cell(row=header_row, column=column).value)
        for column in range(1, ws.max_column + 1)
    ]
    non_empty_headers = [
        (index, header) for index, header in enumerate(headers, start=1) if header
    ]
    diagnostic.non_empty_headers = len(non_empty_headers)
    diagnostic.headers_preview = [header for _, header in non_empty_headers[:30]]
    diagnostic.source_headers = [
        make_simple_candidate(index, header) for index, header in non_empty_headers
    ]

    for index, header in enumerate(headers, start=1):
        if header:
            continue
        has_values = any(
            iter_non_empty_values(
                ws,
                index,
                row_numbers=visible_non_empty_data_rows(ws, header_row),
            )
        )
        if has_values:
            diagnostic.empty_header_columns_with_data.append(get_column_letter(index))
    if diagnostic.empty_header_columns_with_data:
        diagnostic.blockers.append("Colonne(s) avec données mais sans en-tête.")

    seen_headers: dict[str, int] = {}
    for _, header in non_empty_headers:
        seen_headers[header] = seen_headers.get(header, 0) + 1
    diagnostic.duplicate_headers = sorted(
        header for header, count in seen_headers.items() if count > 1
    )
    if diagnostic.duplicate_headers:
        diagnostic.warnings.append(
            "En-têtes sources dupliqués ; la provenance par lettre de colonne permet de les distinguer."
        )

    cleaned_headers: dict[str, int] = {}
    for index, header in non_empty_headers:
        cleaned = clean_indesign_header(f"{get_column_letter(index)}_{header}")
        cleaned_headers[cleaned] = cleaned_headers.get(cleaned, 0) + 1
    diagnostic.cleaned_header_collisions = sorted(
        name for name, count in cleaned_headers.items() if count > 1
    )
    if diagnostic.cleaned_header_collisions:
        diagnostic.blockers.append("Collision après nettoyage des en-têtes InDesign.")

    for index, header in non_empty_headers:
        key = ascii_key(header)
        if ID_HEADER_RE.match(key):
            diagnostic.id_candidates.append(
                make_id_candidate(ws, index, header, header_row)
            )
        if REGION_RE.search(header) or REGION_RE.search(key):
            diagnostic.region_candidates.append(make_simple_candidate(index, header))
        if DEPARTMENT_RE.search(header) or DEPARTMENT_RE.search(key):
            diagnostic.department_candidates.append(
                make_simple_candidate(index, header)
            )
        if DATE_RE.search(header) or DATE_RE.search(key):
            diagnostic.date_candidates.append(make_simple_candidate(index, header))
        if IMAGE_RE.search(header) or IMAGE_RE.search(key):
            diagnostic.image_candidates.append(make_simple_candidate(index, header))
        if SENSITIVE_RE.search(header) or SENSITIVE_RE.search(key):
            diagnostic.sensitive_candidates.append(make_simple_candidate(index, header))

    if not diagnostic.id_candidates:
        diagnostic.blockers.append("Clé primaire dossier non détectée.")
    elif len(diagnostic.id_candidates) > 1:
        diagnostic.blockers.append(
            "Plusieurs colonnes candidates pour la clé primaire dossier."
        )
    else:
        candidate = diagnostic.id_candidates[0]
        if candidate.duplicate_values:
            diagnostic.blockers.append(
                "Valeurs de clé primaire dossier dupliquées dans l'onglet."
            )
        if candidate.blank_values:
            diagnostic.warnings.append(
                "Ligne(s) sans clé primaire dossier à signaler et supprimer à l'export."
            )

    diagnostic.importable = not diagnostic.blockers
    return diagnostic


def diagnose_workbook(
    path: Path,
    limits: ExcelDimensionLimits = DEFAULT_EXCEL_DIMENSION_LIMITS,
) -> WorkbookDiagnostic:
    preflight_sheets = _preflight_workbook_dimensions(path, limits)
    if any(sheet.dimension_limits_exceeded for sheet in preflight_sheets):
        return _build_workbook_diagnostic(path, preflight_sheets)

    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = [diagnose_sheet(sheet, limits) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    return _build_workbook_diagnostic(path, sheets)


def _preflight_workbook_dimensions(
    path: Path,
    limits: ExcelDimensionLimits,
) -> list[SheetDiagnostic]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheets: list[SheetDiagnostic] = []
        for sheet in workbook.worksheets:
            diagnostic = SheetDiagnostic(
                name=sheet.title,
                state=getattr(sheet, "sheet_state", "visible"),
                rows=_dimension_value(sheet.max_row),
                columns=_dimension_value(sheet.max_column),
            )
            try:
                check_worksheet_dimensions(sheet, limits)
            except ExcelDimensionLimitError as exc:
                _block_sheet_for_dimension_limit(diagnostic, exc.violation)
            sheets.append(diagnostic)
        return sheets
    finally:
        workbook.close()


def _build_workbook_diagnostic(
    path: Path,
    sheets: list[SheetDiagnostic],
) -> WorkbookDiagnostic:
    blockers: list[str] = []
    warnings: list[str] = []
    for sheet in sheets:
        for blocker in sheet.blockers:
            blockers.append(f"{sheet.name}: {blocker}")
        for warning in sheet.warnings:
            warnings.append(f"{sheet.name}: {warning}")
    cleaned_header_collisions = workbook_cleaned_header_collisions(sheets)
    if cleaned_header_collisions:
        blockers.append("CSV: Collision après nettoyage des en-têtes InDesign.")

    return WorkbookDiagnostic(
        path=str(path),
        filename=path.name,
        sheet_count=len(sheets),
        importable=not blockers,
        blockers=blockers,
        warnings=warnings,
        sheets=sheets,
        cleaned_header_collisions=cleaned_header_collisions,
    )


def workbook_cleaned_header_collisions(
    sheets: list[SheetDiagnostic],
) -> dict[str, list[str]]:
    cleaned_headers: dict[str, list[str]] = {}
    for sheet in sheets:
        if sheet.ignored or sheet.header_row is None:
            continue
        for source_header in sheet.source_headers:
            if ID_HEADER_RE.match(ascii_key(source_header.header)):
                continue
            cleaned = clean_indesign_header(
                f"{source_header.column}_{source_header.header}"
            )
            cleaned_headers.setdefault(cleaned, []).append(
                f"{sheet.name}!{source_header.column}"
            )
    return {
        cleaned: sources
        for cleaned, sources in sorted(cleaned_headers.items())
        if len(sources) > 1
    }


def _block_sheet_for_dimension_limit(
    diagnostic: SheetDiagnostic,
    violation: ExcelDimensionLimitViolation,
) -> None:
    diagnostic.dimension_limits_exceeded.append(violation.public_details())
    diagnostic.blockers.append("Dimensions Excel hors limites.")
    diagnostic.importable = False


def format_text_report(diagnostics: list[WorkbookDiagnostic]) -> str:
    lines: list[str] = []
    for workbook in diagnostics:
        status = "ACCEPTABLE" if workbook.importable else "REFUSE"
        lines.append(f"## {workbook.filename} - {status}")
        lines.append(f"path: {workbook.path}")
        lines.append(f"sheets: {workbook.sheet_count}")
        if workbook.blockers:
            lines.append("blockers:")
            lines.extend(f"- {item}" for item in workbook.blockers)
        if workbook.warnings:
            lines.append("warnings:")
            lines.extend(f"- {item}" for item in workbook.warnings)
        for sheet in workbook.sheets:
            ignored = " ignored" if sheet.ignored else ""
            sheet_status = "OK" if sheet.importable else "BLOCKED"
            lines.append(
                f"- sheet {sheet.name!r}: {sheet_status}{ignored}, {sheet.rows} rows x {sheet.columns} cols"
            )
            if sheet.ignore_reason:
                lines.append(f"  ignore_reason: {sheet.ignore_reason}")
            lines.append(
                f"  header_row: {sheet.header_row}, non_empty_headers: {sheet.non_empty_headers}"
            )
            if sheet.id_candidates:
                ids = ", ".join(
                    f"{candidate.column}:{candidate.header} "
                    f"(non_empty={candidate.non_empty_values}, unique={candidate.unique_values}, "
                    f"duplicates={candidate.duplicate_values}, blanks={candidate.blank_values})"
                    for candidate in sheet.id_candidates
                )
                lines.append(f"  id_candidates: {ids}")
            else:
                lines.append("  id_candidates: none")
            if sheet.region_candidates:
                lines.append(
                    "  region_candidates: "
                    + ", ".join(
                        f"{candidate.column}:{candidate.header}"
                        for candidate in sheet.region_candidates[:8]
                    )
                )
            if sheet.department_candidates:
                lines.append(
                    "  department_candidates: "
                    + ", ".join(
                        f"{candidate.column}:{candidate.header}"
                        for candidate in sheet.department_candidates[:8]
                    )
                )
            if sheet.date_candidates:
                lines.append(f"  date_candidates: {len(sheet.date_candidates)}")
            if sheet.image_candidates:
                lines.append(
                    "  image_candidates: "
                    + ", ".join(
                        f"{candidate.column}:{candidate.header}"
                        for candidate in sheet.image_candidates[:8]
                    )
                )
            if sheet.duplicate_headers:
                lines.append(
                    "  duplicate_headers: " + ", ".join(sheet.duplicate_headers[:12])
                )
            if sheet.headers_preview:
                lines.append(
                    "  headers_preview: " + " | ".join(sheet.headers_preview[:12])
                )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Diagnose Sircom 2026 Excel input structure."
    )
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument(
        "--json", action="store_true", help="Print JSON instead of a text report."
    )
    args = parser.parse_args(argv)

    diagnostics = [diagnose_workbook(path) for path in args.workbooks]
    if args.json:
        print(
            json.dumps(
                [asdict(diagnostic) for diagnostic in diagnostics],
                ensure_ascii=False,
                indent=2,
            )
        )
    else:
        print(format_text_report(diagnostics), end="")

    return 0 if all(diagnostic.importable for diagnostic in diagnostics) else 1


if __name__ == "__main__":
    raise SystemExit(main())
