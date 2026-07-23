from __future__ import annotations

import uuid
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sircom2026.artifacts import ArtifactStore
from sircom2026.config import Settings
from sircom2026.database import LOT_WRITE_BLOCKED_STATUSES, Repositories
from sircom2026.excel_diagnostic import (
    EXCEL_DIMENSIONS_EXCEEDED_CODE,
    ExcelDimensionLimitError,
    ExcelDimensionLimits,
    check_worksheet_dimensions,
    excel_dimension_limits_from_settings,
)
from sircom2026.invalidation import record_input_change, step_input_fingerprint
from sircom2026.state import complete_step, transition_step
from sircom2026.worker import enqueue_job


ALLOWED_EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
UPLOAD_STEP_KEY = "upload_excel"
DIAGNOSTIC_STEP_KEY = "diagnostic_excel"
UPLOAD_WORKER_ID = "http-upload"
UPLOAD_RULES_VERSION = "upload-excel-v1"


@dataclass(frozen=True)
class ExcelUploadValidation:
    extension: str
    sheet_count: int


@dataclass(frozen=True)
class ExcelUploadResult:
    artifact: dict[str, Any]
    diagnostic_job: dict[str, Any]
    diagnostic_job_created: bool
    invalidated_steps: tuple[str, ...]


class ExcelUploadError(ValueError):
    def __init__(
        self,
        status_code: int,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.code = code
        self.message = message
        self.details = details


def validate_excel_upload(
    *,
    filename: str | None,
    content: bytes,
    max_excel_mb: int,
    dimension_limits: ExcelDimensionLimits,
) -> ExcelUploadValidation:
    extension = Path(filename or "").suffix.lower()
    if extension not in ALLOWED_EXCEL_EXTENSIONS:
        raise ExcelUploadError(
            415,
            "SIRCOM_EXCEL_EXTENSION_UNSUPPORTED",
            "Format Excel non supporté.",
            details={"allowed_extensions": list(ALLOWED_EXCEL_EXTENSIONS)},
        )

    max_bytes = max_excel_mb * 1024 * 1024
    if len(content) > max_bytes:
        raise ExcelUploadError(
            413,
            "SIRCOM_EXCEL_TOO_LARGE",
            "Fichier Excel trop volumineux.",
            details={"max_mb": max_excel_mb, "size_bytes": len(content)},
        )
    if not content:
        raise ExcelUploadError(
            422,
            "SIRCOM_EXCEL_EMPTY",
            "Fichier Excel vide.",
        )

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
        )
    except Exception as exc:
        raise ExcelUploadError(
            422,
            "SIRCOM_EXCEL_UNREADABLE",
            "Archive Excel illisible ou corrompue.",
        ) from exc
    try:
        sheet_count = len(workbook.sheetnames)
        for worksheet in workbook.worksheets:
            check_worksheet_dimensions(worksheet, dimension_limits)
    except ExcelDimensionLimitError as exc:
        raise ExcelUploadError(
            422,
            EXCEL_DIMENSIONS_EXCEEDED_CODE,
            "Classeur Excel hors limites dimensionnelles.",
            details=exc.violation.public_details(),
        ) from exc
    finally:
        workbook.close()

    if sheet_count < 1:
        raise ExcelUploadError(
            422,
            "SIRCOM_EXCEL_UNREADABLE",
            "Archive Excel illisible ou corrompue.",
        )

    return ExcelUploadValidation(extension=extension, sheet_count=sheet_count)


def upload_excel_for_lot(
    repositories: Repositories,
    *,
    settings: Settings,
    lot_id: str,
    filename: str | None,
    content: bytes,
    content_type: str | None,
    idempotency_key: str,
) -> ExcelUploadResult:
    validation = validate_excel_upload(
        filename=filename,
        content=content,
        max_excel_mb=settings.max_excel_mb,
        dimension_limits=excel_dimension_limits_from_settings(settings),
    )
    lot = repositories.lots.get_required(lot_id)
    if lot["status"] in LOT_WRITE_BLOCKED_STATUSES:
        raise ExcelUploadError(
            409,
            "SIRCOM_LOT_NOT_MUTABLE",
            "Lot non modifiable.",
        )
    existing_result = _existing_upload_result(
        repositories,
        lot_id=lot_id,
        idempotency_key=idempotency_key,
    )
    if existing_result is not None:
        return existing_result

    run_id = f"run_{uuid.uuid4().hex}"
    repositories.jobs.cancel_active_for_step(lot_id, UPLOAD_STEP_KEY)
    repositories.artifacts.mark_obsolete_for_steps(
        lot_id=lot_id,
        step_keys=(UPLOAD_STEP_KEY,),
    )
    repositories.steps.prepare_run(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=run_id,
    )
    transition_step(
        repositories,
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        status="en_cours",
        run_id=run_id,
    )
    upload_job = repositories.jobs.create_owned_running(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=run_id,
        idempotency_key=idempotency_key,
        lease_owner=UPLOAD_WORKER_ID,
        lease_seconds=settings.worker_lease_ttl_seconds,
    )

    store = ArtifactStore(
        settings.data_dir,
        pending_ttl_seconds=settings.artifact_pending_ttl_seconds,
    )
    artifact = store.put_temp_then_commit(
        repositories,
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=run_id,
        kind="excel",
        role="source",
        filename=f"source{validation.extension}",
        content=content,
        metadata={
            "content_type": content_type,
            "extension": validation.extension,
            "rules_version": UPLOAD_RULES_VERSION,
            "sheet_count": validation.sheet_count,
        },
        mime_type=EXCEL_MIME_TYPE,
        lease_version=upload_job["lease_version"],
    )
    source_change = record_input_change(
        repositories,
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        input_payload={
            "artifact_id": artifact["id"],
            "extension": validation.extension,
            "rules_version": UPLOAD_RULES_VERSION,
            "sha256": artifact["sha256"],
            "sheet_count": validation.sheet_count,
            "size_bytes": artifact["size_bytes"],
        },
        reason="new_excel",
    )
    finished_job = repositories.jobs.finish_owned(
        job_id=upload_job["id"],
        worker_id=UPLOAD_WORKER_ID,
        run_id=run_id,
        lease_version=upload_job["lease_version"],
        status="succeeded",
    )
    if finished_job is None:
        raise ExcelUploadError(
            409,
            "SIRCOM_UPLOAD_RUN_NOT_CURRENT",
            "Upload Excel interrompu par une exécution plus récente.",
        )
    complete_step(
        repositories,
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=run_id,
    )
    diagnostic_input_fingerprint = step_input_fingerprint(
        repositories,
        lot_id=lot_id,
        step_key=DIAGNOSTIC_STEP_KEY,
    )
    diagnostic = enqueue_job(
        repositories,
        lot_id=lot_id,
        step_key=DIAGNOSTIC_STEP_KEY,
        idempotency_key=f"{DIAGNOSTIC_STEP_KEY}:{artifact['id']}",
        input_fingerprint=diagnostic_input_fingerprint,
    )
    repositories.events.create(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=run_id,
        event_type="excel.uploaded",
        payload={
            "artifact_id": artifact["id"],
            "output_fingerprint": source_change.source_fingerprint,
            "run_id": run_id,
            "size_bytes": artifact["size_bytes"],
            "step_key": UPLOAD_STEP_KEY,
        },
    )
    return ExcelUploadResult(
        artifact=artifact,
        diagnostic_job=diagnostic.job,
        diagnostic_job_created=diagnostic.created,
        invalidated_steps=source_change.invalidated_steps,
    )


def _existing_upload_result(
    repositories: Repositories,
    *,
    lot_id: str,
    idempotency_key: str,
) -> ExcelUploadResult | None:
    existing_job = repositories.jobs.get_by_idempotency_key(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        idempotency_key=idempotency_key,
    )
    if existing_job is None:
        return None
    if existing_job["status"] != "succeeded":
        raise ExcelUploadError(
            409,
            "SIRCOM_UPLOAD_ALREADY_SUBMITTED",
            "Upload Excel déjà soumis.",
        )

    artifact = repositories.artifacts.get_for_step_run_role(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=existing_job["run_id"],
        role="source",
    )
    if artifact is None or artifact["status"] != "committed":
        raise ExcelUploadError(
            409,
            "SIRCOM_UPLOAD_ALREADY_SUBMITTED",
            "Upload Excel déjà soumis.",
        )
    diagnostic_job = repositories.jobs.get_by_idempotency_key(
        lot_id=lot_id,
        step_key=DIAGNOSTIC_STEP_KEY,
        idempotency_key=f"{DIAGNOSTIC_STEP_KEY}:{artifact['id']}",
    )
    if diagnostic_job is None:
        raise ExcelUploadError(
            409,
            "SIRCOM_UPLOAD_ALREADY_SUBMITTED",
            "Upload Excel déjà soumis.",
        )

    return ExcelUploadResult(
        artifact=artifact,
        diagnostic_job=diagnostic_job,
        diagnostic_job_created=False,
        invalidated_steps=(),
    )
