"""Synthetic Excel fixtures for Sircom 2026 diagnostics and app tests."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font


@dataclass(frozen=True)
class SyntheticWorkbookCase:
    name: str
    filename: str
    expected_importable: bool
    description: str


CASES: tuple[SyntheticWorkbookCase, ...] = (
    SyntheticWorkbookCase(
        name="valid_multi_tabs",
        filename="sircom2026-valid-multi-onglets.xlsx",
        expected_importable=True,
        description="Multi-onglets importable, positions id_dossier variées, onglet vide ignoré.",
    ),
    SyntheticWorkbookCase(
        name="missing_id",
        filename="sircom2026-refus-id-manquant.xlsx",
        expected_importable=False,
        description="Onglet utile sans colonne id_dossier.",
    ),
    SyntheticWorkbookCase(
        name="duplicate_id",
        filename="sircom2026-refus-id-duplique.xlsx",
        expected_importable=False,
        description="Doublon de clé primaire id_dossier dans un onglet.",
    ),
    SyntheticWorkbookCase(
        name="ambiguous_id",
        filename="sircom2026-refus-id-ambigu.xlsx",
        expected_importable=False,
        description="Plusieurs colonnes candidates id_dossier dans le même onglet.",
    ),
    SyntheticWorkbookCase(
        name="merged_cells",
        filename="sircom2026-refus-cellules-fusionnees.xlsx",
        expected_importable=False,
        description="Cellules fusionnées détectées.",
    ),
    SyntheticWorkbookCase(
        name="hidden_column",
        filename="sircom2026-refus-colonne-masquee.xlsx",
        expected_importable=False,
        description="Colonne masquée détectée.",
    ),
    SyntheticWorkbookCase(
        name="formula",
        filename="sircom2026-refus-formule.xlsx",
        expected_importable=False,
        description="Formule détectée.",
    ),
    SyntheticWorkbookCase(
        name="multirow_header",
        filename="sircom2026-refus-entete-multiligne.xlsx",
        expected_importable=False,
        description="En-tête détecté hors première ligne.",
    ),
    SyntheticWorkbookCase(
        name="hidden_row",
        filename="sircom2026-alerte-ligne-masquee-ignoree.xlsx",
        expected_importable=True,
        description="Ligne masquée détectée puis ignorée à l'import.",
    ),
    SyntheticWorkbookCase(
        name="hidden_sheet",
        filename="sircom2026-refus-onglet-masque.xlsx",
        expected_importable=False,
        description="Onglet non vide masqué détecté.",
    ),
    SyntheticWorkbookCase(
        name="data_without_header",
        filename="sircom2026-refus-donnees-sans-entete.xlsx",
        expected_importable=False,
        description="Colonne avec données mais sans en-tête.",
    ),
    SyntheticWorkbookCase(
        name="cleaned_header_collision",
        filename="sircom2026-refus-collision-entetes-csv.xlsx",
        expected_importable=False,
        description="Collision de nom CSV après nettoyage multi-onglets.",
    ),
    SyntheticWorkbookCase(
        name="duplicate_source_headers",
        filename="sircom2026-alerte-entetes-sources-dupliques.xlsx",
        expected_importable=True,
        description="Doublon d'en-tête source non bloquant grâce à la provenance.",
    ),
    SyntheticWorkbookCase(
        name="multiple_blockers",
        filename="sircom2026-refus-plusieurs-blocages.xlsx",
        expected_importable=False,
        description="Plusieurs refus stricts détectables en une passe.",
    ),
)


def get_case(name: str) -> SyntheticWorkbookCase:
    for case in CASES:
        if case.name == name:
            return case
    known = ", ".join(case.name for case in CASES)
    raise ValueError(f"Unknown synthetic workbook case {name!r}. Known cases: {known}")


def write_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def save_workbook(workbook: Workbook, output_dir: Path, case_name: str) -> Path:
    case = get_case(case_name)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / case.filename
    workbook.save(path)
    return path


def build_valid_multi_tabs(output_dir: Path) -> Path:
    workbook = Workbook()

    dossiers = workbook.active
    dossiers.title = "Dossiers"
    write_headers(
        dossiers,
        [
            "id_dossier",
            "Région",
            "Département",
            "Nom du produit",
            "Nom de l'entreprise",
            "Photo principale",
            "Date de dépôt",
            "Colonne entierement vide",
        ],
    )
    dossiers.append(
        [
            "ARA07.2026-HGV",
            "Auvergne-Rhône-Alpes",
            "07",
            "Objet de test A",
            "Entreprise synthétique A",
            "Produit A.JPG",
            date(2026, 1, 12),
            None,
        ]
    )
    dossiers.append(
        [
            "BFC21.2026-TEST",
            "Bourgogne-Franche-Comté",
            "21",
            "Objet de test avec\nretour ligne",
            "Entreprise  synthétique  B",
            "produit_b.png",
            "14/02/2026",
            None,
        ]
    )
    dossiers.append(
        [
            "IDF75.2026-ZERO",
            "Île-de-France",
            "75",
            "Objet avec code postal sensible",
            "Entreprise synthétique C",
            "produit-c.webp",
            None,
            None,
        ]
    )
    dossiers.append(
        [
            None,
            "Occitanie",
            "31",
            "Ligne sans identifiant",
            "Entreprise synthétique D",
            None,
            date(2026, 3, 1),
            None,
        ]
    )

    etablissements = workbook.create_sheet("Etablissements")
    write_headers(
        etablissements,
        ["Nom du site", "Code postal", "Dossier ID", "SIRET", "Téléphone"],
    )
    etablissements.append(
        ["Atelier A", "07000", "ARA07.2026-HGV", "12345678901234", "0102030405"]
    )
    etablissements.append(
        ["Atelier B", "21000", "BFC21.2026-TEST", "23456789012345", "0203040506"]
    )
    etablissements.append(
        ["Atelier C", "75001", "IDF75.2026-ZERO", "34567890123456", "0304050607"]
    )

    images = workbook.create_sheet("Images")
    write_headers(images, ["Fichier source", "id dossier", "Légende image"])
    images.append(["Produit A.JPG", "ARA07.2026-HGV", "Photo principale A"])
    images.append(["produit_b.png", "BFC21.2026-TEST", "Photo principale B"])
    images.append(["produit-c.webp", "IDF75.2026-ZERO", "Photo principale C"])

    workbook.create_sheet("Avis")
    return save_workbook(workbook, output_dir, "valid_multi_tabs")


def build_missing_id(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["Région", "Département", "Nom du produit"])
    sheet.append(["Normandie", "14", "Produit sans cle"])
    return save_workbook(workbook, output_dir, "missing_id")


def build_duplicate_id(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit"])
    sheet.append(["DUP-001", "Produit A"])
    sheet.append(["DUP-001", "Produit B"])
    return save_workbook(workbook, output_dir, "duplicate_id")


def build_ambiguous_id(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Dossier ID", "Nom du produit"])
    sheet.append(["AMB-001", "AMB-001", "Produit ambigu"])
    return save_workbook(workbook, output_dir, "ambiguous_id")


def build_merged_cells(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit", "Région"])
    sheet.append(["MER-001", "Produit cellule fusionnee", "Bretagne"])
    sheet.merge_cells("B2:C2")
    return save_workbook(workbook, output_dir, "merged_cells")


def build_hidden_column(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit", "Région"])
    sheet.append(["HID-001", "Produit colonne masquee", "Grand Est"])
    sheet.column_dimensions["C"].hidden = True
    return save_workbook(workbook, output_dir, "hidden_column")


def build_formula(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit", "Libellé calculé"])
    sheet.append(["FOR-001", "Produit formule", None])
    sheet["C2"] = '=A2&" - "&B2'
    return save_workbook(workbook, output_dir, "formula")


def build_multirow_header(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    sheet.append(["Export Sircom 2026", None, None])
    sheet.append(["id_dossier", "Région", "Nom du produit"])
    sheet.append(["MUL-001", "Pays de la Loire", "Produit entete ligne 2"])
    return save_workbook(workbook, output_dir, "multirow_header")


def build_hidden_row(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit", "Région"])
    sheet.append(["ROW-001", "Produit ligne masquee", "Normandie"])
    sheet.row_dimensions[2].hidden = True
    return save_workbook(workbook, output_dir, "hidden_row")


def build_hidden_sheet(output_dir: Path) -> Path:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Notes"
    dossiers = workbook.create_sheet("Dossiers")
    write_headers(dossiers, ["id_dossier", "Nom du produit"])
    dossiers.append(["SHEET-001", "Produit onglet masque"])
    dossiers.sheet_state = "hidden"
    return save_workbook(workbook, output_dir, "hidden_sheet")


def build_data_without_header(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", None, "Nom du produit"])
    sheet.append(["NOHEAD-001", "Valeur sans en-tete", "Produit sans entete"])
    return save_workbook(workbook, output_dir, "data_without_header")


def build_cleaned_header_collision(output_dir: Path) -> Path:
    workbook = Workbook()
    dossiers = workbook.active
    dossiers.title = "Dossiers"
    write_headers(dossiers, ["id_dossier", "Nom du produit"])
    dossiers.append(["COL-001", "Produit dossier"])

    produits = workbook.create_sheet("Produits")
    write_headers(produits, ["id_dossier", "Nom du produit"])
    produits.append(["COL-001", "Produit collision"])
    return save_workbook(workbook, output_dir, "cleaned_header_collision")


def build_duplicate_source_headers(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["id_dossier", "Nom du produit", "Nom du produit"])
    sheet.append(["DUPHEAD-001", "Produit A", "Produit B"])
    return save_workbook(workbook, output_dir, "duplicate_source_headers")


def build_multiple_blockers(output_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dossiers"
    write_headers(sheet, ["Nom du produit", "Région", "Libellé calculé"])
    sheet.append(["Produit sale", "Bretagne", None])
    sheet["C2"] = '=A2&" - "&B2'
    sheet.column_dimensions["B"].hidden = True
    sheet.row_dimensions[2].hidden = True
    sheet.merge_cells("A2:B2")
    return save_workbook(workbook, output_dir, "multiple_blockers")


BUILDERS = {
    "valid_multi_tabs": build_valid_multi_tabs,
    "missing_id": build_missing_id,
    "duplicate_id": build_duplicate_id,
    "ambiguous_id": build_ambiguous_id,
    "merged_cells": build_merged_cells,
    "hidden_column": build_hidden_column,
    "formula": build_formula,
    "multirow_header": build_multirow_header,
    "hidden_row": build_hidden_row,
    "hidden_sheet": build_hidden_sheet,
    "data_without_header": build_data_without_header,
    "cleaned_header_collision": build_cleaned_header_collision,
    "duplicate_source_headers": build_duplicate_source_headers,
    "multiple_blockers": build_multiple_blockers,
}


def create_synthetic_excels(
    output_dir: Path, case_names: Iterable[str] | None = None
) -> dict[str, Path]:
    selected = (
        list(case_names) if case_names is not None else [case.name for case in CASES]
    )
    created: dict[str, Path] = {}
    for name in selected:
        builder = BUILDERS[name]
        created[name] = builder(output_dir)
    return created


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Create synthetic Sircom 2026 Excel workbooks."
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("tests/fixtures/synthetic-excels"),
        help="Directory where synthetic .xlsx files are written.",
    )
    parser.add_argument(
        "--case",
        action="append",
        choices=[case.name for case in CASES],
        help="Generate only this case. Can be passed several times.",
    )
    args = parser.parse_args(argv)

    created = create_synthetic_excels(args.output_dir, args.case)
    for case in CASES:
        if case.name not in created:
            continue
        print(
            f"{case.name}: {created[case.name]} expected_importable={case.expected_importable}"
        )
        print(f"  {case.description}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
