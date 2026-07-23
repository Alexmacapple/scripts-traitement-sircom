from __future__ import annotations

import json
import math
import posixpath
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

from sircom2026.artifacts import ArtifactStore, ArtifactUnavailableError
from sircom2026.config import Settings
from sircom2026.database import Repositories
from sircom2026.invalidation import fingerprint_payload
from sircom2026.mapping import MAPPING_STEP_KEY
from sircom2026.state import record_problem
from sircom2026.worker import JobResult, WorkerJobContext, WorkerLeaseLost


UPLOAD_STEP_KEY = "upload_excel"
FUSION_STEP_KEY = "fusion_multi_onglets"
FUSION_ARTIFACT_KIND = "json"
FUSION_ARTIFACT_ROLE = "result"
FUSION_RULES_VERSION = "flat-merge-v1"
FUSION_MIME_TYPE = "application/json"
FUSION_SCHEMA_VERSION = 1
NORMALIZATION_STEP_KEY = "normalisation_contenu"
NORMALIZATION_ARTIFACT_KIND = "json"
NORMALIZATION_ARTIFACT_ROLE = "result"
NORMALIZATION_RULES_VERSION = "content-normalisation-v1"
NORMALIZATION_MIME_TYPE = "application/json"
NORMALIZATION_SCHEMA_VERSION = 1
SPECIAL_CSV_NAMES = {"id_dossier", "imageid", "@pathimg"}
SENSITIVE_TEXT_ROLES = {
    "id_dossier",
    "siret",
    "telephone",
    "code_postal",
    "departement",
    "code_administratif",
}
_HORIZONTAL_SPACE_RE = re.compile(r"[ \t\f\v]+")
_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


@dataclass(frozen=True)
class CurrentJsonArtifact:
    artifact: dict[str, Any]
    payload: dict[str, Any]


@dataclass(frozen=True)
class FlatMergeResult:
    payload: dict[str, Any]
    removed_rows_without_id_count: int
    removed_empty_columns_count: int
    duplicate_id_locations: tuple[dict[str, Any], ...] = ()
    missing_sheet_locations: tuple[dict[str, Any], ...] = ()

    @property
    def has_blocker(self) -> bool:
        return bool(self.duplicate_id_locations or self.missing_sheet_locations)

    @property
    def has_warnings(self) -> bool:
        return self.removed_rows_without_id_count > 0


@dataclass(frozen=True)
class ContentNormalizationResult:
    payload: dict[str, Any]
    date_issues_count: int
    invalid_dates_count: int
    missing_dates_count: int
    removed_empty_columns_count: int

    @property
    def has_warnings(self) -> bool:
        return self.date_issues_count > 0


def run_flat_merge_job(context: WorkerJobContext, *, settings: Settings) -> JobResult:
    store = ArtifactStore(
        settings.data_dir,
        pending_ttl_seconds=settings.artifact_pending_ttl_seconds,
    )
    context.set_progress(1, 4)
    with context.database.transaction() as repositories:
        _require_current_lease(repositories, context)
        mapping = _current_validated_mapping(repositories, store, context.lot_id)
        source_artifact = _current_excel_source_artifact(repositories, context.lot_id)
        if mapping is None:
            _record_missing_mapping_problem(repositories, context)
            return JobResult(final_step_status="bloque")
        if source_artifact is None:
            _record_missing_source_problem(repositories, context)
            return JobResult(final_step_status="bloque")
        try:
            readable_source = store.open_for_read(
                repositories,
                lot_id=context.lot_id,
                artifact_id=source_artifact["id"],
            )
        except (ArtifactUnavailableError, KeyError, ValueError):
            _record_missing_source_problem(repositories, context)
            return JobResult(final_step_status="bloque")

    context.set_progress(2, 4)
    merge_result = build_flat_merge(readable_source.path, mapping.payload)
    if merge_result.has_blocker:
        with context.database.transaction() as repositories:
            _require_current_lease(repositories, context)
            repositories.problems.mark_open_obsolete_for_steps(
                lot_id=context.lot_id,
                step_keys=(FUSION_STEP_KEY,),
            )
            for location in merge_result.missing_sheet_locations:
                record_problem(
                    repositories,
                    lot_id=context.lot_id,
                    step_key=FUSION_STEP_KEY,
                    run_id=context.run_id,
                    severity="bloquant",
                    code="SIRCOM_FUSION_SOURCE_SHEET_MISSING",
                    title="Onglet source introuvable",
                    cause="La fusion ne trouve pas un onglet référencé par le mapping validé.",
                    action="Relancer le diagnostic et valider un mapping compatible avec l'Excel courant.",
                    location=location,
                    technical={"rows_count": 0},
                )
            for location in merge_result.duplicate_id_locations:
                record_problem(
                    repositories,
                    lot_id=context.lot_id,
                    step_key=FUSION_STEP_KEY,
                    run_id=context.run_id,
                    severity="bloquant",
                    code="SIRCOM_FUSION_DUPLICATE_ID",
                    title="Doublons id_dossier",
                    cause="Un onglet contient plusieurs lignes avec le même id_dossier.",
                    action="Corriger les doublons id_dossier, puis relancer le diagnostic et le mapping.",
                    location=location,
                    technical={"rows_count": 1},
                )
        return JobResult(final_step_status="bloque")

    content = json.dumps(
        merge_result.payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")

    context.set_progress(3, 4)
    with context.database.transaction() as repositories:
        _require_current_lease(repositories, context)
        repositories.problems.mark_open_obsolete_for_steps(
            lot_id=context.lot_id,
            step_keys=(FUSION_STEP_KEY,),
        )
        repositories.artifacts.mark_obsolete_for_steps(
            lot_id=context.lot_id,
            step_keys=(FUSION_STEP_KEY,),
        )
        artifact = store.put_temp_then_commit(
            repositories,
            lot_id=context.lot_id,
            step_key=FUSION_STEP_KEY,
            run_id=context.run_id,
            kind=FUSION_ARTIFACT_KIND,
            role=FUSION_ARTIFACT_ROLE,
            filename="fusion-multi-onglets.json",
            content=content,
            metadata={
                "columns_count": merge_result.payload["columns_count"],
                "mapping_artifact_id": mapping.artifact["id"],
                "removed_empty_columns_count": merge_result.removed_empty_columns_count,
                "removed_rows_without_id_count": merge_result.removed_rows_without_id_count,
                "rows_count": merge_result.payload["rows_count"],
                "rules_version": FUSION_RULES_VERSION,
                "schema_version": FUSION_SCHEMA_VERSION,
                "source_artifact_id": source_artifact["id"],
            },
            mime_type=FUSION_MIME_TYPE,
            lease_version=context.leased_job.lease_version,
        )
        if merge_result.removed_rows_without_id_count:
            record_problem(
                repositories,
                lot_id=context.lot_id,
                step_key=FUSION_STEP_KEY,
                run_id=context.run_id,
                severity="alerte",
                code="SIRCOM_FUSION_ROWS_WITHOUT_ID_REMOVED",
                title="Lignes sans id_dossier supprimées",
                cause="Certaines lignes sans id_dossier ont été ignorées pendant la fusion.",
                action="Compléter ces identifiants dans l'Excel source ou accepter leur suppression.",
                technical={"rows_removed": merge_result.removed_rows_without_id_count},
            )
        output_fingerprint = fingerprint_payload(
            {
                "artifact_sha256": artifact["sha256"],
                "fusion_artifact_id": artifact["id"],
                "kind": "flat_merge",
                "mapping_artifact_id": mapping.artifact["id"],
                "rules_version": FUSION_RULES_VERSION,
                "schema_version": FUSION_SCHEMA_VERSION,
                "source_artifact_id": source_artifact["id"],
            }
        )
        repositories.events.create(
            lot_id=context.lot_id,
            step_key=FUSION_STEP_KEY,
            run_id=context.run_id,
            event_type="fusion.completed",
            payload={
                "artifact_id": artifact["id"],
                "columns_count": merge_result.payload["columns_count"],
                "rows_count": merge_result.payload["rows_count"],
                "rows_removed": merge_result.removed_rows_without_id_count,
                "status": "termine_avec_alertes"
                if merge_result.has_warnings
                else "termine",
                "step_key": FUSION_STEP_KEY,
            },
        )

    context.set_progress(4, 4)
    return JobResult(
        with_warnings=merge_result.has_warnings,
        output_fingerprint=output_fingerprint,
        enqueue_next_steps=(NORMALIZATION_STEP_KEY,),
    )


def run_content_normalization_job(
    context: WorkerJobContext, *, settings: Settings
) -> JobResult:
    store = ArtifactStore(
        settings.data_dir,
        pending_ttl_seconds=settings.artifact_pending_ttl_seconds,
    )
    context.set_progress(1, 4)
    with context.database.transaction() as repositories:
        _require_current_lease(repositories, context)
        fusion = _current_json_artifact(
            repositories,
            store,
            lot_id=context.lot_id,
            step_key=FUSION_STEP_KEY,
            role=FUSION_ARTIFACT_ROLE,
            ready_statuses=("termine", "termine_avec_alertes"),
        )
        if fusion is None:
            _record_missing_fusion_problem(repositories, context)
            return JobResult(final_step_status="bloque")

    context.set_progress(2, 4)
    normalization_result = normalize_flat_merge(
        fusion.payload,
        source_fusion_artifact_id=fusion.artifact["id"],
    )
    content = json.dumps(
        normalization_result.payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")

    context.set_progress(3, 4)
    with context.database.transaction() as repositories:
        _require_current_lease(repositories, context)
        repositories.problems.mark_open_obsolete_for_steps(
            lot_id=context.lot_id,
            step_keys=(NORMALIZATION_STEP_KEY,),
        )
        repositories.artifacts.mark_obsolete_for_steps(
            lot_id=context.lot_id,
            step_keys=(NORMALIZATION_STEP_KEY,),
        )
        artifact = store.put_temp_then_commit(
            repositories,
            lot_id=context.lot_id,
            step_key=NORMALIZATION_STEP_KEY,
            run_id=context.run_id,
            kind=NORMALIZATION_ARTIFACT_KIND,
            role=NORMALIZATION_ARTIFACT_ROLE,
            filename="normalisation-contenu.json",
            content=content,
            metadata={
                "columns_count": normalization_result.payload["columns_count"],
                "date_issues_count": normalization_result.date_issues_count,
                "invalid_dates_count": normalization_result.invalid_dates_count,
                "missing_dates_count": normalization_result.missing_dates_count,
                "removed_empty_columns_count": (
                    normalization_result.removed_empty_columns_count
                ),
                "rows_count": normalization_result.payload["rows_count"],
                "rules_version": NORMALIZATION_RULES_VERSION,
                "schema_version": NORMALIZATION_SCHEMA_VERSION,
                "source_fusion_artifact_id": fusion.artifact["id"],
            },
            mime_type=NORMALIZATION_MIME_TYPE,
            lease_version=context.leased_job.lease_version,
        )
        if normalization_result.date_issues_count:
            record_problem(
                repositories,
                lot_id=context.lot_id,
                step_key=NORMALIZATION_STEP_KEY,
                run_id=context.run_id,
                severity="alerte",
                code="SIRCOM_NORMALIZATION_DATE_VALUES_INVALID",
                title="Dates invalides ou absentes",
                cause="Certaines valeurs de colonnes date ne peuvent pas être exportées comme dates valides.",
                action="Corriger les dates dans l'Excel source ou accepter une valeur #N/A à l'export.",
                technical={
                    "date_issues_count": normalization_result.date_issues_count,
                    "invalid_dates_count": normalization_result.invalid_dates_count,
                    "missing_dates_count": normalization_result.missing_dates_count,
                },
            )
        output_fingerprint = fingerprint_payload(
            {
                "artifact_sha256": artifact["sha256"],
                "kind": "content_normalization",
                "normalization_artifact_id": artifact["id"],
                "rules_version": NORMALIZATION_RULES_VERSION,
                "schema_version": NORMALIZATION_SCHEMA_VERSION,
                "source_fusion_artifact_id": fusion.artifact["id"],
            }
        )
        repositories.events.create(
            lot_id=context.lot_id,
            step_key=NORMALIZATION_STEP_KEY,
            run_id=context.run_id,
            event_type="normalization.completed",
            payload={
                "artifact_id": artifact["id"],
                "columns_count": normalization_result.payload["columns_count"],
                "rows_count": normalization_result.payload["rows_count"],
                "status": (
                    "termine_avec_alertes"
                    if normalization_result.has_warnings
                    else "termine"
                ),
                "step_key": NORMALIZATION_STEP_KEY,
                "warning_code": (
                    "SIRCOM_NORMALIZATION_DATE_VALUES_INVALID"
                    if normalization_result.has_warnings
                    else None
                ),
            },
        )

    context.set_progress(4, 4)
    return JobResult(
        with_warnings=normalization_result.has_warnings,
        output_fingerprint=output_fingerprint,
        enqueue_next_steps=("verification_csv_indesign",),
        require_next_validations=("tri_region_departement",),
    )


def build_flat_merge(workbook_path, mapping: dict[str, Any]) -> FlatMergeResult:
    exported_columns = _ordered_exported_columns(mapping)
    source_columns = [
        column
        for column in exported_columns
        if not bool(column.get("system")) and column.get("logical_role") != "id_dossier"
    ]
    id_columns = _id_columns_by_sheet(mapping)
    columns_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for column in source_columns:
        source_sheet = str(column["source_sheet"])
        columns_by_sheet.setdefault(source_sheet, []).append(column)

    rows_by_id: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    removed_rows: list[dict[str, Any]] = []
    duplicate_locations: list[dict[str, Any]] = []
    missing_sheet_locations: list[dict[str, Any]] = []
    source_rows_count = 0
    source_rank = 0
    hidden_rows_by_sheet = _hidden_rows_by_sheet(workbook_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_info in _ordered_sheets(mapping):
            sheet_name = str(sheet_info["name"])
            if sheet_name not in workbook.sheetnames:
                missing_sheet_locations.append({"onglet": sheet_name})
                continue
            id_column = id_columns.get(sheet_name)
            if id_column is None:
                continue
            worksheet = workbook[sheet_name]
            header_row = _positive_int(sheet_info.get("header_row"), default=1)
            hidden_rows = hidden_rows_by_sheet.get(sheet_name, set())
            seen_ids_in_sheet: set[str] = set()
            for row_number, row_values in enumerate(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                if row_number in hidden_rows:
                    continue
                if _row_values_are_empty(row_values):
                    continue
                source_rows_count += 1
                id_value = _row_value_at(
                    row_values,
                    id_column["source_column_index"],
                )
                id_dossier = _identifier_text(id_value)
                if not id_dossier:
                    removed_rows.append(
                        {"source_sheet": sheet_name, "row_number": row_number}
                    )
                    continue
                if id_dossier in seen_ids_in_sheet:
                    duplicate_locations.append(
                        {
                            "onglet": sheet_name,
                            "ligne": row_number,
                            "colonne": id_column["source_column_letter"],
                        }
                    )
                    continue
                seen_ids_in_sheet.add(id_dossier)

                source_rank += 1
                row = rows_by_id.get(id_dossier)
                if row is None:
                    row = {
                        "source_rank": source_rank,
                        "id_dossier": id_dossier,
                        "values": {
                            column["csv_name"]: "" for column in exported_columns
                        },
                    }
                    row["values"]["id_dossier"] = id_dossier
                    rows_by_id[id_dossier] = row
                    rows.append(row)

                for column in columns_by_sheet.get(sheet_name, []):
                    value = _row_value_at(row_values, column["source_column_index"])
                    row["values"][column["csv_name"]] = _json_cell_value(value)
    finally:
        workbook.close()

    if duplicate_locations or missing_sheet_locations:
        return FlatMergeResult(
            payload={},
            removed_rows_without_id_count=len(removed_rows),
            removed_empty_columns_count=0,
            duplicate_id_locations=tuple(duplicate_locations[:100]),
            missing_sheet_locations=tuple(missing_sheet_locations[:100]),
        )

    kept_columns, removed_empty_columns = _remove_empty_columns(exported_columns, rows)
    kept_names = {column["csv_name"] for column in kept_columns}
    for row in rows:
        row["values"] = {
            column["csv_name"]: row["values"].get(column["csv_name"], "")
            for column in kept_columns
        }

    payload = {
        "schema_version": FUSION_SCHEMA_VERSION,
        "rules_version": FUSION_RULES_VERSION,
        "source_diagnostic_artifact_id": mapping.get("source_diagnostic_artifact_id"),
        "structural_fingerprint": mapping.get("structural_fingerprint"),
        "rows_count": len(rows),
        "columns_count": len(kept_columns),
        "source_rows_count": source_rows_count,
        "removed_rows_without_id_count": len(removed_rows),
        "removed_empty_columns_count": len(removed_empty_columns),
        "columns": [_public_column(column) for column in kept_columns],
        "rows": rows,
        "removed_rows_without_id": removed_rows[:100],
        "removed_empty_columns": [
            _public_column(column)
            for column in removed_empty_columns
            if column["csv_name"] not in kept_names
        ],
    }
    return FlatMergeResult(
        payload=payload,
        removed_rows_without_id_count=len(removed_rows),
        removed_empty_columns_count=len(removed_empty_columns),
    )


def _hidden_rows_by_sheet(workbook_path: Any) -> dict[str, set[int]]:
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            workbook_rels = _workbook_relationships(archive)
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            sheets_el = workbook_root.find(_xlsx_tag("sheets"))
            if sheets_el is None:
                return {}

            hidden_by_sheet: dict[str, set[int]] = {}
            for sheet_el in sheets_el.findall(_xlsx_tag("sheet")):
                sheet_name = str(sheet_el.attrib.get("name") or "")
                relationship_id = sheet_el.attrib.get(f"{{{_OFFICE_REL_NS}}}id")
                sheet_path = workbook_rels.get(str(relationship_id or ""))
                if not sheet_name or not sheet_path:
                    continue
                hidden_by_sheet[sheet_name] = _hidden_rows_in_sheet_xml(
                    archive,
                    sheet_path,
                )
            return hidden_by_sheet
    except (ET.ParseError, KeyError, OSError, ValueError, zipfile.BadZipFile):
        return {}


def _workbook_relationships(archive: zipfile.ZipFile) -> dict[str, str]:
    relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships: dict[str, str] = {}
    for relationship in relationships_root.findall(_relationship_tag("Relationship")):
        relationship_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target")
        if relationship_id and target:
            relationships[relationship_id] = _resolve_xlsx_target(
                "xl/_rels/workbook.xml.rels",
                target,
            )
    return relationships


def _hidden_rows_in_sheet_xml(
    archive: zipfile.ZipFile,
    sheet_path: str,
) -> set[int]:
    sheet_root = ET.fromstring(archive.read(sheet_path))
    sheet_data = sheet_root.find(_xlsx_tag("sheetData"))
    if sheet_data is None:
        return set()
    hidden_rows: set[int] = set()
    for row_el in sheet_data.findall(_xlsx_tag("row")):
        if row_el.attrib.get("hidden") != "1":
            continue
        try:
            hidden_rows.add(int(row_el.attrib.get("r") or "0"))
        except ValueError:
            continue
    hidden_rows.discard(0)
    return hidden_rows


def _resolve_xlsx_target(relationship_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    if "/_rels/" in relationship_path:
        source_folder = relationship_path.split("/_rels/", 1)[0]
    else:
        source_folder = posixpath.dirname(relationship_path)
    return posixpath.normpath(posixpath.join(source_folder, target))


def _xlsx_tag(name: str) -> str:
    return f"{{{_SPREADSHEET_NS}}}{name}"


def _relationship_tag(name: str) -> str:
    return f"{{{_RELATIONSHIPS_NS}}}{name}"


def normalize_flat_merge(
    fusion_payload: dict[str, Any],
    *,
    source_fusion_artifact_id: str | None = None,
) -> ContentNormalizationResult:
    columns = [dict(column) for column in fusion_payload.get("columns", [])]
    date_issues: list[dict[str, Any]] = []
    invalid_dates_count = 0
    missing_dates_count = 0
    rows: list[dict[str, Any]] = []

    for source_row in fusion_payload.get("rows", []):
        if not isinstance(source_row, dict):
            continue
        source_values = (
            source_row.get("values")
            if isinstance(source_row.get("values"), dict)
            else {}
        )
        values: dict[str, Any] = {}
        for column in columns:
            csv_name = str(column["csv_name"])
            normalized_value, date_issue = _normalize_cell_value(
                source_values.get(csv_name, ""),
                column,
            )
            values[csv_name] = normalized_value
            if date_issue:
                if date_issue == "invalid":
                    invalid_dates_count += 1
                else:
                    missing_dates_count += 1
                date_issues.append(
                    {
                        "csv_name": csv_name,
                        "issue": date_issue,
                        "source_column_letter": column.get("source_column_letter"),
                        "source_header": column.get("source_header"),
                        "source_rank": source_row.get("source_rank"),
                        "source_sheet": column.get("source_sheet"),
                    }
                )

        id_dossier = str(values.get("id_dossier", "") or "")
        rows.append(
            {
                "source_rank": source_row.get("source_rank"),
                "id_dossier": id_dossier,
                "values": values,
            }
        )

    kept_columns, removed_empty_columns = _remove_empty_columns(columns, rows)
    for row in rows:
        row["values"] = {
            column["csv_name"]: row["values"].get(column["csv_name"], "")
            for column in kept_columns
        }

    date_issues_count = invalid_dates_count + missing_dates_count
    payload = {
        "schema_version": NORMALIZATION_SCHEMA_VERSION,
        "rules_version": NORMALIZATION_RULES_VERSION,
        "source_fusion_artifact_id": source_fusion_artifact_id,
        "source_fusion_rules_version": fusion_payload.get("rules_version"),
        "structural_fingerprint": fusion_payload.get("structural_fingerprint"),
        "rows_count": len(rows),
        "columns_count": len(kept_columns),
        "date_issues_count": date_issues_count,
        "invalid_dates_count": invalid_dates_count,
        "missing_dates_count": missing_dates_count,
        "removed_empty_columns_count": len(removed_empty_columns),
        "upstream_removed_empty_columns_count": fusion_payload.get(
            "removed_empty_columns_count",
            0,
        ),
        "upstream_removed_rows_without_id_count": fusion_payload.get(
            "removed_rows_without_id_count",
            0,
        ),
        "upstream_removed_empty_columns": fusion_payload.get(
            "removed_empty_columns", []
        ),
        "upstream_removed_rows_without_id": fusion_payload.get(
            "removed_rows_without_id", []
        ),
        "columns": [_public_column(column) for column in kept_columns],
        "rows": rows,
        "date_issues": date_issues[:100],
        "removed_empty_columns": [
            _public_column(column) for column in removed_empty_columns
        ],
    }
    return ContentNormalizationResult(
        payload=payload,
        date_issues_count=date_issues_count,
        invalid_dates_count=invalid_dates_count,
        missing_dates_count=missing_dates_count,
        removed_empty_columns_count=len(removed_empty_columns),
    )


def _ordered_exported_columns(mapping: dict[str, Any]) -> list[dict[str, Any]]:
    columns = [
        dict(column)
        for column in mapping.get("columns", [])
        if isinstance(column, dict) and column.get("status") == "exporte"
    ]
    return sorted(
        columns, key=lambda column: int(column.get("output_position") or 999_999)
    )


def _ordered_sheets(mapping: dict[str, Any]) -> list[dict[str, Any]]:
    sheets = [sheet for sheet in mapping.get("sheets", []) if isinstance(sheet, dict)]
    return sheets


def _id_columns_by_sheet(mapping: dict[str, Any]) -> dict[str, dict[str, Any]]:
    id_columns: dict[str, dict[str, Any]] = {}
    for column in mapping.get("columns", []):
        if (
            isinstance(column, dict)
            and not bool(column.get("system"))
            and column.get("logical_role") == "id_dossier"
            and column.get("source_sheet")
        ):
            id_columns[str(column["source_sheet"])] = dict(column)
    return id_columns


def _remove_empty_columns(
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    kept: list[dict[str, Any]] = []
    removed: list[dict[str, Any]] = []
    for column in columns:
        csv_name = str(column["csv_name"])
        if csv_name in SPECIAL_CSV_NAMES:
            kept.append(column)
            continue
        if all(_cell_is_empty(row["values"].get(csv_name, "")) for row in rows):
            removed.append(column)
        else:
            kept.append(column)
    return kept, removed


def _public_column(column: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": column["id"],
        "system": bool(column.get("system")),
        "source_sheet": column.get("source_sheet"),
        "source_column_letter": column.get("source_column_letter"),
        "source_header": column.get("source_header"),
        "logical_role": column.get("logical_role"),
        "csv_name": column["csv_name"],
        "output_position": column.get("output_position"),
    }


def _json_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _normalize_cell_value(value: Any, column: dict[str, Any]) -> tuple[Any, str | None]:
    role = str(column.get("logical_role") or "")
    if role == "date":
        return _normalize_date_value(value)
    if _is_blank_value(value):
        return "", None
    if role in SENSITIVE_TEXT_ROLES:
        return _normalize_text_value(_text_from_value(value)), None
    if isinstance(value, str):
        return _normalize_text_value(value), None
    if isinstance(value, float) and math.isnan(value):
        return "", None
    if isinstance(value, Decimal):
        return float(value), None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(), None
    return value, None


def _normalize_date_value(value: Any) -> tuple[str, str | None]:
    if _is_blank_value(value):
        return "", "missing"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y"), None
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y"), None
    text = _normalize_text_value(_text_from_value(value))
    parsed = _parse_date_text(text)
    if parsed is None:
        return "", "invalid"
    return parsed.strftime("%d/%m/%Y"), None


def _parse_date_text(text: str) -> date | None:
    if not text:
        return None
    iso_text = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(iso_text).date()
    except ValueError:
        pass
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass
    for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _normalize_text_value(value: str) -> str:
    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    parts = [
        _HORIZONTAL_SPACE_RE.sub(" ", part).strip() for part in normalized.split("\n")
    ]
    return "<br>".join(parts).strip()


def _text_from_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _row_values_are_empty(row_values: tuple[Any, ...]) -> bool:
    return all(_is_blank_value(value) for value in row_values)


def _row_value_at(row_values: tuple[Any, ...], column_index: Any) -> Any:
    try:
        index = int(column_index) - 1
    except (TypeError, ValueError):
        return None
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def _identifier_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_is_empty(value: Any) -> bool:
    return value is None or value == ""


def _positive_int(value: Any, *, default: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _current_validated_mapping(
    repositories: Repositories,
    store: ArtifactStore,
    lot_id: str,
) -> CurrentJsonArtifact | None:
    mapping_step = repositories.steps.get_by_lot_key(lot_id, MAPPING_STEP_KEY)
    if (
        mapping_step is None
        or not mapping_step["current_run_id"]
        or mapping_step["status"] not in {"termine", "termine_avec_alertes"}
    ):
        return None
    artifact = repositories.artifacts.get_for_step_run_role(
        lot_id=lot_id,
        step_key=MAPPING_STEP_KEY,
        run_id=mapping_step["current_run_id"],
        role="validated",
    )
    if artifact is None or artifact["status"] != "committed":
        return None
    try:
        readable = store.open_for_read(
            repositories,
            lot_id=lot_id,
            artifact_id=artifact["id"],
        )
        payload = json.loads(readable.path.read_text(encoding="utf-8"))
    except (
        ArtifactUnavailableError,
        OSError,
        json.JSONDecodeError,
        KeyError,
        ValueError,
    ):
        return None
    if not isinstance(payload, dict):
        return None
    return CurrentJsonArtifact(artifact=artifact, payload=payload)


def _current_json_artifact(
    repositories: Repositories,
    store: ArtifactStore,
    *,
    lot_id: str,
    step_key: str,
    role: str,
    ready_statuses: tuple[str, ...],
) -> CurrentJsonArtifact | None:
    step = repositories.steps.get_by_lot_key(lot_id, step_key)
    if (
        step is None
        or not step["current_run_id"]
        or step["status"] not in ready_statuses
    ):
        return None
    artifact = repositories.artifacts.get_for_step_run_role(
        lot_id=lot_id,
        step_key=step_key,
        run_id=step["current_run_id"],
        role=role,
    )
    if artifact is None or artifact["status"] != "committed":
        return None
    try:
        readable = store.open_for_read(
            repositories,
            lot_id=lot_id,
            artifact_id=artifact["id"],
        )
        payload = json.loads(readable.path.read_text(encoding="utf-8"))
    except (
        ArtifactUnavailableError,
        OSError,
        json.JSONDecodeError,
        KeyError,
        ValueError,
    ):
        return None
    if not isinstance(payload, dict):
        return None
    return CurrentJsonArtifact(artifact=artifact, payload=payload)


def _current_excel_source_artifact(
    repositories: Repositories,
    lot_id: str,
) -> dict[str, Any] | None:
    upload_step = repositories.steps.get_by_lot_key(lot_id, UPLOAD_STEP_KEY)
    if upload_step is None or not upload_step["current_run_id"]:
        return None
    artifact = repositories.artifacts.get_for_step_run_role(
        lot_id=lot_id,
        step_key=UPLOAD_STEP_KEY,
        run_id=upload_step["current_run_id"],
        role="source",
    )
    if artifact is None or artifact["status"] != "committed":
        return None
    return artifact


def _require_current_lease(
    repositories: Repositories,
    context: WorkerJobContext,
) -> None:
    if (
        repositories.jobs.get_committable_by_run(
            lot_id=context.lot_id,
            step_key=context.step_key,
            run_id=context.run_id,
            lease_version=context.leased_job.lease_version,
            expected_input_fingerprint=context.leased_job.input_fingerprint,
        )
        is None
    ):
        raise WorkerLeaseLost("Worker lease is no longer current.")


def _record_missing_mapping_problem(
    repositories: Repositories,
    context: WorkerJobContext,
) -> None:
    repositories.problems.mark_open_obsolete_for_steps(
        lot_id=context.lot_id,
        step_keys=(FUSION_STEP_KEY,),
    )
    record_problem(
        repositories,
        lot_id=context.lot_id,
        step_key=FUSION_STEP_KEY,
        run_id=context.run_id,
        severity="bloquant",
        code="SIRCOM_FUSION_MAPPING_NOT_VALIDATED",
        title="Mapping validé introuvable",
        cause="La fusion ne trouve pas de snapshot de mapping validé courant.",
        action="Valider le mapping avant de relancer la fusion.",
    )


def _record_missing_source_problem(
    repositories: Repositories,
    context: WorkerJobContext,
) -> None:
    repositories.problems.mark_open_obsolete_for_steps(
        lot_id=context.lot_id,
        step_keys=(FUSION_STEP_KEY,),
    )
    record_problem(
        repositories,
        lot_id=context.lot_id,
        step_key=FUSION_STEP_KEY,
        run_id=context.run_id,
        severity="bloquant",
        code="SIRCOM_FUSION_SOURCE_EXCEL_MISSING",
        title="Excel source introuvable",
        cause="La fusion ne trouve pas l'artefact Excel source courant.",
        action="Déposer à nouveau le fichier Excel, puis relancer le diagnostic et le mapping.",
    )


def _record_missing_fusion_problem(
    repositories: Repositories,
    context: WorkerJobContext,
) -> None:
    repositories.problems.mark_open_obsolete_for_steps(
        lot_id=context.lot_id,
        step_keys=(NORMALIZATION_STEP_KEY,),
    )
    record_problem(
        repositories,
        lot_id=context.lot_id,
        step_key=NORMALIZATION_STEP_KEY,
        run_id=context.run_id,
        severity="bloquant",
        code="SIRCOM_NORMALIZATION_FUSION_NOT_READY",
        title="Fusion multi-onglets introuvable",
        cause="La normalisation ne trouve pas de table fusionnée courante.",
        action="Relancer la fusion multi-onglets avant la normalisation.",
    )
