from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import sircom2026.excel_diagnostic as excel_diagnostic
from sircom2026.excel_diagnostic import (
    ExcelDimensionLimits,
    clean_indesign_header,
    diagnose_sheet,
    diagnose_workbook,
)
from sircom2026.synthetic_excels import CASES, create_synthetic_excels


REAL_SIRCOM1 = Path("livrables-miweb-2025/Sircom1.xlsx")
REAL_SIRCOM2 = Path("livrables-miweb-2025/Sircom2.xlsx")


class ExcelDiagnosticTest(unittest.TestCase):
    def test_synthetic_cases_have_expected_import_status(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = create_synthetic_excels(Path(tmpdir))

            for case in CASES:
                with self.subTest(case=case.name):
                    diagnostic = diagnose_workbook(paths[case.name])
                    self.assertEqual(
                        diagnostic.importable,
                        case.expected_importable,
                        diagnostic.blockers,
                    )

    def test_valid_fixture_exercises_multi_tab_id_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = create_synthetic_excels(Path(tmpdir), ["valid_multi_tabs"])
            diagnostic = diagnose_workbook(paths["valid_multi_tabs"])

        self.assertTrue(diagnostic.importable, diagnostic.blockers)
        sheets = {sheet.name: sheet for sheet in diagnostic.sheets}
        self.assertEqual(sheets["Dossiers"].id_candidates[0].column, "A")
        self.assertEqual(sheets["Etablissements"].id_candidates[0].column, "C")
        self.assertEqual(sheets["Images"].id_candidates[0].column, "B")
        self.assertTrue(sheets["Avis"].ignored)
        self.assertTrue(
            any("sans id_dossier" in warning for warning in diagnostic.warnings),
            diagnostic.warnings,
        )
        self.assertTrue(
            any("Onglet vide ignoré." in warning for warning in diagnostic.warnings),
            diagnostic.warnings,
        )

    def test_2025_header_cleaning_rule_keeps_excel_letter_prefix(self) -> None:
        self.assertEqual(clean_indesign_header("B_ID"), "b_id")
        self.assertEqual(
            clean_indesign_header("CE_Une (seule) photo du produit candidat"),
            "ce_uneseul",
        )
        self.assertEqual(clean_indesign_header("AC_Département"), "ac_departe")

    def test_refusal_reasons_are_actionable_for_common_bad_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = create_synthetic_excels(
                Path(tmpdir),
                [
                    "missing_id",
                    "duplicate_id",
                    "ambiguous_id",
                    "merged_cells",
                    "hidden_column",
                    "hidden_row",
                    "hidden_sheet",
                    "formula",
                    "multirow_header",
                    "data_without_header",
                    "cleaned_header_collision",
                ],
            )
            expected_fragments = {
                "missing_id": "Colonne id_dossier non détectée.",
                "duplicate_id": "Valeurs id_dossier dupliquées",
                "ambiguous_id": "Plusieurs colonnes id_dossier candidates.",
                "merged_cells": "cellule(s) fusionnée(s).",
                "hidden_column": "colonne(s) masquée(s).",
                "hidden_row": "ligne(s) masquée(s).",
                "hidden_sheet": "Onglet masqué.",
                "formula": "Formules détectées.",
                "multirow_header": "En-tête détecté hors première ligne.",
                "data_without_header": "Colonne(s) avec données mais sans en-tête.",
                "cleaned_header_collision": "Collision après nettoyage des en-têtes InDesign.",
            }

            for case_name, fragment in expected_fragments.items():
                with self.subTest(case=case_name):
                    diagnostic = diagnose_workbook(paths[case_name])
                    self.assertFalse(diagnostic.importable)
                    self.assertIn(fragment, " ".join(diagnostic.blockers))

    def test_duplicate_source_headers_are_non_blocking_when_provenance_disambiguates(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = create_synthetic_excels(Path(tmpdir), ["duplicate_source_headers"])
            diagnostic = diagnose_workbook(paths["duplicate_source_headers"])

        self.assertTrue(diagnostic.importable, diagnostic.blockers)
        self.assertIn("En-têtes sources dupliqués", " ".join(diagnostic.warnings))

    def test_refused_workbook_reports_multiple_detectable_reasons_in_one_pass(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = create_synthetic_excels(Path(tmpdir), ["multiple_blockers"])
            diagnostic = diagnose_workbook(paths["multiple_blockers"])

        blockers = " ".join(diagnostic.blockers)
        self.assertFalse(diagnostic.importable)
        self.assertIn("colonne(s) masquée(s).", blockers)
        self.assertIn("ligne(s) masquée(s).", blockers)
        self.assertIn("cellule(s) fusionnée(s).", blockers)
        self.assertIn("Formules détectées.", blockers)
        self.assertIn("Colonne id_dossier non détectée.", blockers)

    def test_cell_scan_limit_does_not_trust_declared_dimensions_only(self) -> None:
        class FakeCell:
            def __init__(self, value: str) -> None:
                self.value = value

        class FakeWorksheet:
            title = "Produits"
            sheet_state = "visible"
            max_row = 1
            max_column = 1

            def iter_rows(self):
                yield [FakeCell("id_dossier")]
                yield [FakeCell("DOSSIER-1")]

        diagnostic = diagnose_sheet(
            FakeWorksheet(),
            limits=ExcelDimensionLimits(max_rows=10, max_columns=10, max_cells=1),
        )

        self.assertFalse(diagnostic.importable)
        self.assertEqual(
            diagnostic.dimension_limits_exceeded[0]["limit_exceeded"], "max_cells"
        )
        self.assertEqual(diagnostic.dimension_limits_exceeded[0]["observed"], 2)

    def test_direct_workbook_diagnostic_blocks_dimensions_before_full_load(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "oversized.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Produits"
            sheet.append(["id_dossier", "nom"])
            sheet.append(["DOSSIER-1", "Produit 1"])
            sheet.append(["DOSSIER-2", "Produit 2"])
            workbook.save(workbook_path)
            workbook.close()

            original_load_workbook = excel_diagnostic.load_workbook

            def guarded_load_workbook(*args, **kwargs):
                self.assertIs(
                    kwargs.get("read_only"),
                    True,
                    "diagnose_workbook must preflight dimensions before full load",
                )
                return original_load_workbook(*args, **kwargs)

            with patch(
                "sircom2026.excel_diagnostic.load_workbook",
                side_effect=guarded_load_workbook,
            ):
                diagnostic = diagnose_workbook(
                    workbook_path,
                    limits=ExcelDimensionLimits(
                        max_rows=2,
                        max_columns=10,
                        max_cells=100,
                    ),
                )

        self.assertFalse(diagnostic.importable)
        self.assertEqual(diagnostic.sheet_count, 1)
        self.assertEqual(diagnostic.sheets[0].name, "Produits")
        self.assertEqual(
            diagnostic.sheets[0].dimension_limits_exceeded[0]["limit_exceeded"],
            "max_rows",
        )
        self.assertEqual(
            diagnostic.sheets[0].dimension_limits_exceeded[0]["observed"],
            3,
        )
        self.assertIn("Produits: Dimensions Excel hors limites.", diagnostic.blockers)

    def test_local_2024_2025_inputs_when_available(self) -> None:
        if not REAL_SIRCOM1.exists() or not REAL_SIRCOM2.exists():
            self.skipTest(
                "Local Sircom1.xlsx and Sircom2.xlsx fixtures are not present."
            )

        legacy_limits = ExcelDimensionLimits(
            max_rows=2_000_000,
            max_columns=20_000,
            max_cells=2_000_000_000,
        )
        sircom1 = diagnose_workbook(REAL_SIRCOM1, limits=legacy_limits)
        sircom2 = diagnose_workbook(REAL_SIRCOM2, limits=legacy_limits)

        self.assertFalse(sircom1.importable)
        self.assertIn("colonne(s) masquée(s)", " ".join(sircom1.blockers))
        self.assertIn("Formules détectées.", " ".join(sircom1.blockers))
        self.assertTrue(sircom2.importable, sircom2.blockers)


if __name__ == "__main__":
    unittest.main()
