from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest import mock

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from sircom2026.app import create_app
from sircom2026.config import load_settings
from sircom2026.database import Database
from sircom2026.transform import build_flat_merge
from sircom2026.worker_runner import run_worker_once


def make_settings(tmpdir: Path):
    return load_settings(
        {
            "SIRCOM_DATA_DIR": str(tmpdir / "data"),
            "SIRCOM_SQLITE_PATH": str(tmpdir / "data" / "sircom.sqlite3"),
            "SIRCOM_DISK_FREE_MIN_MB": "0",
        }
    )


def excel_file(path: Path) -> dict[str, tuple[str, bytes, str]]:
    return {
        "file": (
            path.name,
            path.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


def create_merge_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    dossiers = workbook.active
    dossiers.title = "Dossiers"
    dossiers.append(["id_dossier", "Nom produit", "Colonne vide"])
    dossiers.append(["ONLY-A", "Produit A", None])
    dossiers.append(["COMMON", "Produit commun", None])
    dossiers.append([None, "Ligne sans identifiant", None])
    dossiers.append(["HIDDEN-A", "Produit masque", None])
    dossiers.row_dimensions[5].hidden = True

    complement = workbook.create_sheet("Complement")
    complement.append(["id_dossier", "Région", "Note vide"])
    complement.append(["COMMON", "Bretagne", None])
    complement.append(["ONLY-B", "Occitanie", None])

    workbook.save(path)
    workbook.close()


def direct_merge_mapping() -> dict[str, Any]:
    return {
        "source_diagnostic_artifact_id": "diagnostic-test",
        "structural_fingerprint": "fingerprint-test",
        "sheets": [
            {"name": "Dossiers", "header_row": 1},
            {"name": "Complement", "header_row": 1},
        ],
        "columns": [
            {
                "id": "Dossiers!A",
                "status": "exporte",
                "system": False,
                "source_sheet": "Dossiers",
                "source_column_index": 1,
                "source_column_letter": "A",
                "source_header": "id_dossier",
                "logical_role": "id_dossier",
                "csv_name": "id_dossier",
                "output_position": 1,
            },
            {
                "id": "system:imageid",
                "status": "exporte",
                "system": True,
                "source_sheet": None,
                "source_column_index": None,
                "source_column_letter": None,
                "source_header": "Image InDesign générée",
                "logical_role": "nom_image_source",
                "csv_name": "imageid",
                "output_position": 2,
            },
            {
                "id": "system:@pathimg",
                "status": "exporte",
                "system": True,
                "source_sheet": None,
                "source_column_index": None,
                "source_column_letter": None,
                "source_header": "Chemin image InDesign",
                "logical_role": "nom_image_source",
                "csv_name": "@pathimg",
                "output_position": 3,
            },
            {
                "id": "Dossiers!B",
                "status": "exporte",
                "system": False,
                "source_sheet": "Dossiers",
                "source_column_index": 2,
                "source_column_letter": "B",
                "source_header": "Nom produit",
                "logical_role": "texte",
                "csv_name": "b_nomprodu",
                "output_position": 4,
            },
            {
                "id": "Dossiers!C",
                "status": "exporte",
                "system": False,
                "source_sheet": "Dossiers",
                "source_column_index": 3,
                "source_column_letter": "C",
                "source_header": "Colonne vide",
                "logical_role": "texte",
                "csv_name": "c_colonnev",
                "output_position": 5,
            },
            {
                "id": "Complement!A",
                "status": "exporte",
                "system": False,
                "source_sheet": "Complement",
                "source_column_index": 1,
                "source_column_letter": "A",
                "source_header": "id_dossier",
                "logical_role": "id_dossier",
                "csv_name": "id_dossier",
                "output_position": 6,
            },
            {
                "id": "Complement!B",
                "status": "exporte",
                "system": False,
                "source_sheet": "Complement",
                "source_column_index": 2,
                "source_column_letter": "B",
                "source_header": "Région",
                "logical_role": "texte",
                "csv_name": "b_region",
                "output_position": 7,
            },
            {
                "id": "Complement!C",
                "status": "exporte",
                "system": False,
                "source_sheet": "Complement",
                "source_column_index": 3,
                "source_column_letter": "C",
                "source_header": "Note vide",
                "logical_role": "texte",
                "csv_name": "c_notevide",
                "output_position": 8,
            },
        ],
    }


def prepare_importable_lot(
    client: TestClient,
    settings,
    excel_path: Path,
    *,
    title: str,
    key: str,
) -> str:
    lot_id = client.post("/api/lots", json={"title": title}).json()["lot"]["id"]
    upload = client.post(
        f"/api/lots/{lot_id}/excel",
        files=excel_file(excel_path),
        headers={"X-Idempotency-Key": f"upload-{key}"},
    )
    worker_result = run_worker_once(settings=settings)

    if upload.status_code != 202:
        raise AssertionError(upload.text)
    if worker_result.outcome != "succeeded":
        raise AssertionError(worker_result)
    return lot_id


def mapping_submission(mapping: dict[str, Any]) -> dict[str, Any]:
    return {
        "structural_fingerprint": mapping["structural_fingerprint"],
        "columns": [
            {
                "id": column["id"],
                "status": column["status"],
                "csv_name": column["csv_name"],
                "logical_role": column["logical_role"],
                "suppression_reason": column["suppression_reason"],
            }
            for column in mapping["columns"]
        ],
    }


def step_status(lot: dict[str, object], step_key: str) -> str:
    for step in lot["steps"]:
        if step["key"] == step_key:
            return str(step["status"])
    raise AssertionError(f"Missing step {step_key}.")


def download_fusion_payload(
    client: TestClient, settings, lot_id: str
) -> dict[str, Any]:
    database = Database(settings.sqlite_path)
    with database.session() as repositories:
        step = repositories.steps.get_by_lot_key(lot_id, "fusion_multi_onglets")
        if step is None or not step["current_run_id"]:
            raise AssertionError("Fusion step has no current run.")
        artifact = repositories.artifacts.get_for_step_run_role(
            lot_id=lot_id,
            step_key="fusion_multi_onglets",
            run_id=step["current_run_id"],
            role="result",
        )
        if artifact is None:
            raise AssertionError("Fusion result artifact is missing.")

    response = client.get(f"/api/lots/{lot_id}/downloads/{artifact['id']}")
    if response.status_code != 200:
        raise AssertionError(response.text)
    return response.json()


class FusionWorkerTest(unittest.TestCase):
    def test_flat_merge_streams_read_only_rows_without_random_cell_lookups(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "fixtures" / "fusion.xlsx"
            create_merge_workbook(workbook_path)

            with mock.patch.object(
                ReadOnlyWorksheet,
                "cell",
                side_effect=AssertionError(
                    "build_flat_merge must stream read-only rows with iter_rows"
                ),
            ):
                result = build_flat_merge(workbook_path, direct_merge_mapping())

        payload = result.payload
        self.assertEqual(payload["rows_count"], 3)
        self.assertEqual(payload["source_rows_count"], 5)
        self.assertEqual(payload["removed_rows_without_id_count"], 1)
        self.assertEqual(payload["removed_empty_columns_count"], 2)
        row_by_id = {row["id_dossier"]: row["values"] for row in payload["rows"]}
        self.assertEqual(list(row_by_id), ["ONLY-A", "COMMON", "ONLY-B"])
        self.assertNotIn("HIDDEN-A", row_by_id)
        self.assertEqual(row_by_id["COMMON"]["b_region"], "Bretagne")

    def test_mapping_validation_schedules_and_worker_merges_multi_sheet_rows(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            workbook_path = tmpdir / "fixtures" / "fusion.xlsx"
            create_merge_workbook(workbook_path)

            settings = make_settings(tmpdir)
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                workbook_path,
                title="Lot fusion",
                key="fusion",
            )
            mapping = client.get(f"/api/lots/{lot_id}/mapping").json()["mapping"]

            validate = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "fusion-mapping-validate"},
            )
            worker_result = run_worker_once(settings=settings)
            lot = client.get(f"/api/lots/{lot_id}").json()["lot"]
            payload = download_fusion_payload(client, settings, lot_id)

        self.assertEqual(validate.status_code, 200, validate.text)
        self.assertEqual(
            step_status(validate.json()["lot"], "fusion_multi_onglets"), "pret"
        )
        self.assertEqual(worker_result.outcome, "succeeded")
        self.assertEqual(worker_result.step_key, "fusion_multi_onglets")
        self.assertEqual(
            step_status(lot, "fusion_multi_onglets"), "termine_avec_alertes"
        )

        self.assertEqual(payload["schema_version"], 1)
        self.assertEqual(payload["rules_version"], "flat-merge-v1")
        self.assertEqual(payload["rows_count"], 3)
        self.assertEqual(payload["removed_rows_without_id_count"], 1)
        self.assertEqual(payload["removed_empty_columns_count"], 2)

        csv_names = [column["csv_name"] for column in payload["columns"]]
        self.assertEqual(csv_names.count("id_dossier"), 1)
        self.assertEqual(csv_names[:3], ["id_dossier", "imageid", "@pathimg"])
        self.assertIn("b_nomprodu", csv_names)
        self.assertIn("b_region", csv_names)
        self.assertNotIn("c_colonnev", csv_names)
        self.assertNotIn("c_notevide", csv_names)

        row_by_id = {row["id_dossier"]: row["values"] for row in payload["rows"]}
        self.assertEqual(list(row_by_id), ["ONLY-A", "COMMON", "ONLY-B"])
        self.assertNotIn("HIDDEN-A", row_by_id)
        self.assertEqual(row_by_id["ONLY-A"]["b_nomprodu"], "Produit A")
        self.assertEqual(row_by_id["ONLY-A"]["b_region"], "")
        self.assertEqual(row_by_id["COMMON"]["b_nomprodu"], "Produit commun")
        self.assertEqual(row_by_id["COMMON"]["b_region"], "Bretagne")
        self.assertEqual(row_by_id["ONLY-B"]["b_nomprodu"], "")
        self.assertEqual(row_by_id["ONLY-B"]["b_region"], "Occitanie")
        self.assertEqual(row_by_id["COMMON"]["imageid"], "")
        self.assertEqual(row_by_id["COMMON"]["@pathimg"], "")
        self.assertNotIn(
            "Ligne sans identifiant", str(payload["removed_rows_without_id"])
        )


if __name__ == "__main__":
    unittest.main()
