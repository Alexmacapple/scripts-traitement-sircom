from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook

import sircom2026.mapping as mapping_module
from sircom2026.app import create_app
from sircom2026.config import load_settings
from sircom2026.mapping import (
    FUSION_STEP_KEY,
    MAPPING_LOGICAL_ROLES,
    MAPPING_MIME_TYPE,
    MAPPING_RULES_VERSION,
    MAPPING_SCHEMA_VERSION,
    MAPPING_STEP_KEY,
    MAPPING_STATUS_VALUES,
    SYSTEM_COLUMN_IDS,
    MappingError,
    MappingOperationResult,
    PersistedMappingSnapshot,
)
from sircom2026.synthetic_excels import create_synthetic_excels
from sircom2026.worker_runner import run_worker_once


def make_settings(tmpdir: Path):
    return load_settings(
        {
            "SIRCOM_DATA_DIR": str(tmpdir / "data"),
            "SIRCOM_SQLITE_PATH": str(tmpdir / "data" / "sircom.sqlite3"),
            "SIRCOM_DISK_FREE_MIN_MB": "0",
        }
    )


def excel_file(path: Path) -> dict[str, tuple[str, bytes, str]]:
    return {
        "file": (
            path.name,
            path.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


def create_late_header_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BDD"
    sheet.cell(row=40, column=1, value="ID")
    sheet.cell(row=40, column=2, value="Région")
    sheet.cell(row=40, column=3, value="Nom du produit")
    sheet.cell(row=41, column=1, value="DOSSIER-001")
    sheet.cell(row=41, column=2, value="Bretagne")
    sheet.cell(row=41, column=3, value="Produit tardif")
    workbook.save(path)
    workbook.close()


def prepare_importable_lot(
    client: TestClient,
    settings,
    excel_path: Path,
    *,
    title: str,
    key: str,
) -> str:
    lot_id = client.post("/api/lots", json={"title": title}).json()["lot"]["id"]
    upload = client.post(
        f"/api/lots/{lot_id}/excel",
        files=excel_file(excel_path),
        headers={"X-Idempotency-Key": f"upload-{key}"},
    )
    worker_result = run_worker_once(settings=settings)

    if upload.status_code != 202:
        raise AssertionError(upload.text)
    worker_results = [worker_result]
    for _ in range(10):
        lot = client.get(f"/api/lots/{lot_id}").json()["lot"]
        if step_status(lot, "diagnostic_excel") in {"termine", "termine_avec_alertes"}:
            return lot_id
        worker_results.append(run_worker_once(settings=settings))
    raise AssertionError(worker_results)


class MappingPublicContractTest(unittest.TestCase):
    def test_public_mapping_module_contract_entries_are_stable(self) -> None:
        public_callables = {
            "get_mapping_payload",
            "build_default_mapping_from_current_diagnostic",
            "save_mapping_draft",
            "validate_mapping",
            "save_profile_from_validated_mapping",
            "apply_profile_as_draft",
            "read_current_mapping_artifact",
            "mapping_from_submission",
            "mapping_validation_errors",
            "apply_profile_to_default_mapping",
            "profile_compatibility",
            "profile_from_mapping",
            "MappingProfileStore",
        }

        for name in public_callables:
            with self.subTest(name=name):
                self.assertTrue(callable(getattr(mapping_module, name, None)))

        self.assertTrue(issubclass(MappingError, ValueError))
        self.assertEqual(
            tuple(MappingOperationResult.__dataclass_fields__),
            ("mapping", "artifact", "lot", "invalidated_steps"),
        )
        self.assertEqual(
            tuple(PersistedMappingSnapshot.__dataclass_fields__),
            ("artifact", "created"),
        )
        self.assertEqual(MAPPING_SCHEMA_VERSION, 1)
        self.assertEqual(MAPPING_RULES_VERSION, "mapping-v1")
        self.assertEqual(MAPPING_STEP_KEY, "mapping")
        self.assertEqual(FUSION_STEP_KEY, "fusion_multi_onglets")
        self.assertEqual(MAPPING_MIME_TYPE, "application/json")
        self.assertEqual(MAPPING_STATUS_VALUES, {"exporte", "supprime"})
        self.assertEqual(SYSTEM_COLUMN_IDS, ("system:imageid", "system:@pathimg"))
        self.assertEqual(
            MAPPING_LOGICAL_ROLES,
            {
                "id_dossier",
                "date",
                "region",
                "departement",
                "nom_image_source",
                "siret",
                "telephone",
                "code_postal",
                "code_administratif",
                "texte",
            },
        )
        error = MappingError(
            422,
            "SIRCOM_MAPPING_CONTRACT_TEST",
            "Erreur de contrat mapping.",
            details={"field": "mapping"},
        )
        self.assertEqual(error.status_code, 422)
        self.assertEqual(error.code, "SIRCOM_MAPPING_CONTRACT_TEST")
        self.assertEqual(error.message, "Erreur de contrat mapping.")
        self.assertEqual(error.details, {"field": "mapping"})


class MappingApiTest(unittest.TestCase):
    def test_default_mapping_selects_useful_columns_and_keeps_structural_provenance(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures", ["valid_multi_tabs"]
            )
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot mapping",
                key="mapping-default",
            )

            response = client.get(f"/api/lots/{lot_id}/mapping")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        mapping = payload["mapping"]
        columns = mapping["columns"]
        self.assertEqual(mapping["source"], "default")
        self.assertEqual(mapping["rules_version"], "mapping-v1")
        self.assertEqual(mapping["schema_version"], 1)
        self.assertEqual(len(mapping["structural_fingerprint"]), 64)
        self.assertTrue(mapping["source_diagnostic_artifact_id"])
        self.assertEqual(
            {
                (sheet["name"], sheet["header_row"], sheet["columns_count"])
                for sheet in mapping["sheets"]
            },
            {
                ("Dossiers", 1, 8),
                ("Etablissements", 1, 5),
                ("Images", 1, 3),
            },
        )
        self.assertNotIn("Objet de test", str(payload))
        self.assertNotIn(str(tmpdir), str(payload))

        source_columns = [column for column in columns if not column["system"]]
        self.assertEqual(len(source_columns), 16)
        required_column_fields = {
            "id",
            "system",
            "source_sheet",
            "source_column_index",
            "source_column_letter",
            "source_header",
            "logical_role",
            "status",
            "csv_name",
            "default_csv_name",
            "suppression_reason",
            "output_position",
            "locked",
        }
        self.assertTrue(
            all(required_column_fields <= set(column) for column in columns)
        )
        self.assertTrue(
            all(
                column["source_sheet"]
                and column["source_column_letter"]
                and column["source_header"]
                for column in source_columns
            )
        )

        exported_csv_names = [
            column["csv_name"] for column in columns if column["status"] == "exporte"
        ]
        self.assertEqual(exported_csv_names.count("id_dossier"), 1)
        id_position = exported_csv_names.index("id_dossier")
        self.assertEqual(
            exported_csv_names[id_position + 1 : id_position + 3],
            ["imageid", "@pathimg"],
        )
        image_columns = [
            column for column in columns if column["id"] in SYSTEM_COLUMN_IDS
        ]
        self.assertEqual(
            [column["id"] for column in image_columns], list(SYSTEM_COLUMN_IDS)
        )
        self.assertTrue(all(column["system"] for column in image_columns))
        self.assertTrue(all(column["locked"] for column in image_columns))
        self.assertEqual(
            [column["csv_name"] for column in image_columns],
            ["imageid", "@pathimg"],
        )
        self.assertIn("d_nomdupro", exported_csv_names)
        self.assertIn("b_region", exported_csv_names)
        self.assertIn("g_datedede", exported_csv_names)
        for column in columns:
            if column["status"] != "exporte":
                continue
            if column["csv_name"] in {"id_dossier", "imageid", "@pathimg"}:
                continue
            self.assertLessEqual(len(column["csv_name"]), 10, column)

        duplicate_id_columns = [
            column
            for column in columns
            if column["logical_role"] == "id_dossier" and column["status"] == "supprime"
        ]
        self.assertEqual(len(duplicate_id_columns), 2)
        self.assertTrue(
            all(column["suppression_reason"] for column in duplicate_id_columns)
        )
        empty_column = find_column(mapping, "Dossiers", "H", "Colonne entierement vide")
        self.assertEqual(empty_column["status"], "exporte")
        self.assertEqual(empty_column["csv_name"], "h_colonnee")
        self.assertLessEqual(len(empty_column["csv_name"]), 10)
        source_image_column = find_column(mapping, "Dossiers", "F", "Photo principale")
        self.assertEqual(source_image_column["logical_role"], "nom_image_source")
        self.assertEqual(source_image_column["status"], "exporte")

    def test_default_mapping_keeps_auto_detected_late_header_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            workbook_path = tmpdir / "fixtures" / "late-header.xlsx"
            create_late_header_workbook(workbook_path)
            settings = make_settings(tmpdir)
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                workbook_path,
                title="Lot entete ligne 40",
                key="late-header",
            )

            response = client.get(f"/api/lots/{lot_id}/mapping")

        self.assertEqual(response.status_code, 200)
        mapping = response.json()["mapping"]
        self.assertEqual(
            [
                (sheet["name"], sheet["header_row"], sheet["columns_count"])
                for sheet in mapping["sheets"]
            ],
            [("BDD", 40, 3)],
        )
        id_column = find_column(mapping, "BDD", "A", "ID")
        self.assertEqual(id_column["logical_role"], "id_dossier")
        self.assertEqual(id_column["csv_name"], "id_dossier")

    def test_mapping_draft_validation_and_profile_persistence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures", ["valid_multi_tabs"]
            )
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot mapping validation",
                key="mapping-validation",
            )
            mapping = client.get(f"/api/lots/{lot_id}/mapping").json()["mapping"]
            product_column = find_column(mapping, "Dossiers", "D", "Nom du produit")
            product_column["csv_name"] = "d_produit"

            draft = client.post(
                f"/api/lots/{lot_id}/mapping/draft",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-draft-1"},
            )
            validate = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-validate-1"},
            )
            validate_replay = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-validate-1"},
            )
            product_column["csv_name"] = "d_autre"
            validate_reused_key = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-validate-1"},
            )
            profile = client.post(
                f"/api/lots/{lot_id}/mapping/profile",
                json={"name": "Profil Sircom test"},
            )
            mapping_view = client.get(f"/?lot_id={lot_id}&view=mapping")
            detail = client.get(f"/api/lots/{lot_id}").json()["lot"]
            validated_mapping = client.get(f"/api/lots/{lot_id}/mapping").json()[
                "mapping"
            ]

        self.assertEqual(draft.status_code, 200, draft.text)
        self.assertEqual(draft.json()["mapping"]["source"], "draft")
        self.assertEqual(step_status(draft.json()["lot"], "mapping"), "action_requise")
        self.assertEqual(validate.status_code, 200, validate.text)
        self.assertEqual(validate.json()["mapping"]["source"], "validated")
        self.assertIn("fusion_multi_onglets", validate.json()["invalidated_steps"])
        self.assertEqual(validate_replay.status_code, 200, validate_replay.text)
        self.assertEqual(validate_replay.json()["invalidated_steps"], [])
        self.assertEqual(
            validate_replay.json()["artifact"]["id"],
            validate.json()["artifact"]["id"],
        )
        self.assertEqual(validate_reused_key.status_code, 409)
        self.assertEqual(
            validate_reused_key.json()["error"]["code"],
            "SIRCOM_MAPPING_IDEMPOTENCY_REUSED",
        )
        self.assertEqual(step_status(detail, "mapping"), "termine")
        self.assertEqual(
            validated_mapping["columns"][product_column["output_position"] - 1][
                "csv_name"
            ],
            "d_produit",
        )

        self.assertEqual(profile.status_code, 201, profile.text)
        saved_profile = profile.json()["profile"]
        self.assertEqual(saved_profile["version"], 1)
        self.assertEqual(
            saved_profile["structural_fingerprint"], mapping["structural_fingerprint"]
        )
        self.assertEqual(saved_profile["sheets"], mapping["sheets"])
        self.assertTrue(saved_profile["headers"])
        self.assertTrue(saved_profile["letters"])
        self.assertTrue(saved_profile["logical_roles"])
        self.assertTrue(saved_profile["columns"])
        self.assertTrue(
            {
                "id",
                "status",
                "csv_name",
                "logical_role",
                "suppression_reason",
            }
            <= set(saved_profile["columns"][0])
        )
        self.assertTrue(saved_profile["last_used_at"])
        self.assertNotIn("Objet de test", str(saved_profile))
        self.assertNotIn(str(tmpdir), str(saved_profile))
        self.assertEqual(mapping_view.status_code, 200, mapping_view.text)
        self.assertNotIn("Nom du profil", mapping_view.text)
        self.assertNotIn("Sauvegarder un profil", mapping_view.text)
        self.assertNotIn('id="mapping-profile-form"', mapping_view.text)

    def test_compatible_profile_is_only_applied_as_draft_after_user_action(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures", ["valid_multi_tabs"]
            )
            client = TestClient(create_app(settings))
            first_lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot profil source",
                key="profile-source",
            )
            first_mapping = client.get(f"/api/lots/{first_lot_id}/mapping").json()[
                "mapping"
            ]
            find_column(first_mapping, "Dossiers", "D", "Nom du produit")[
                "csv_name"
            ] = "d_produit"
            validate = client.post(
                f"/api/lots/{first_lot_id}/mapping/validate",
                json=mapping_submission(first_mapping),
                headers={"X-Idempotency-Key": "profile-source-validate"},
            )
            profile = client.post(
                f"/api/lots/{first_lot_id}/mapping/profile",
                json={"name": "Profil compatible"},
            ).json()["profile"]

            second_lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot profil cible",
                key="profile-target",
            )
            before_apply = client.get(f"/api/lots/{second_lot_id}/mapping").json()
            apply_profile = client.post(
                f"/api/lots/{second_lot_id}/mapping/profile-draft",
                json={"profile_id": profile["id"]},
                headers={"X-Idempotency-Key": "profile-target-draft"},
            )
            after_apply = client.get(f"/api/lots/{second_lot_id}/mapping").json()

        self.assertEqual(validate.status_code, 200, validate.text)
        self.assertEqual(before_apply["mapping"]["source"], "default")
        self.assertEqual(
            find_column(before_apply["mapping"], "Dossiers", "D", "Nom du produit")[
                "csv_name"
            ],
            "d_nomdupro",
        )
        self.assertEqual(len(before_apply["profiles"]["compatible"]), 1)
        self.assertEqual(before_apply["profiles"]["compatible"][0]["id"], profile["id"])
        self.assertEqual(apply_profile.status_code, 200, apply_profile.text)
        self.assertEqual(after_apply["mapping"]["source"], "profile_draft")
        self.assertEqual(
            step_status(apply_profile.json()["lot"], "mapping"), "action_requise"
        )
        self.assertEqual(
            find_column(after_apply["mapping"], "Dossiers", "D", "Nom du produit")[
                "csv_name"
            ],
            "d_produit",
        )

    def test_incompatible_profile_is_reported_and_refused(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures",
                ["valid_multi_tabs", "duplicate_source_headers"],
            )
            client = TestClient(create_app(settings))
            source_lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot profil incompatible source",
                key="profile-incompatible-source",
            )
            source_mapping = client.get(f"/api/lots/{source_lot_id}/mapping").json()[
                "mapping"
            ]
            client.post(
                f"/api/lots/{source_lot_id}/mapping/validate",
                json=mapping_submission(source_mapping),
                headers={"X-Idempotency-Key": "profile-incompatible-validate"},
            )
            profile = client.post(
                f"/api/lots/{source_lot_id}/mapping/profile",
                json={"name": "Profil incompatible"},
            ).json()["profile"]
            target_lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["duplicate_source_headers"],
                title="Lot profil incompatible cible",
                key="profile-incompatible-target",
            )

            profile_view = client.get(f"/api/lots/{target_lot_id}/mapping").json()[
                "profiles"
            ]
            apply_profile = client.post(
                f"/api/lots/{target_lot_id}/mapping/profile-draft",
                json={"profile_id": profile["id"]},
                headers={"X-Idempotency-Key": "profile-incompatible-draft"},
            )

        self.assertEqual(len(profile_view["compatible"]), 0)
        self.assertEqual(len(profile_view["incompatible"]), 1)
        self.assertEqual(profile_view["incompatible"][0]["id"], profile["id"])
        self.assertEqual(apply_profile.status_code, 409)
        self.assertEqual(
            apply_profile.json()["error"]["code"],
            "SIRCOM_MAPPING_PROFILE_INCOMPATIBLE",
        )

    def test_validation_blocks_collisions_and_default_cleaning_handles_accents(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            workbook_path = tmpdir / "fixtures" / "accent.xlsx"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Dossiers"
            sheet.append(["id_dossier", "Élégance supérieure longue", "Region"])
            sheet.append(["ACC-001", "valeur", "Occitanie"])
            complement = workbook.create_sheet("Complement")
            complement.append(["id_dossier", "Autre libellé"])
            complement.append(["ACC-001", "valeur complementaire"])
            workbook.save(workbook_path)
            workbook.close()

            settings = make_settings(tmpdir)
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                workbook_path,
                title="Lot collision mapping",
                key="mapping-collision",
            )
            mapping = client.get(f"/api/lots/{lot_id}/mapping").json()["mapping"]
            accented = find_column(
                mapping, "Dossiers", "B", "Élégance supérieure longue"
            )
            complement_column = find_column(mapping, "Complement", "B", "Autre libellé")
            accented["csv_name"] = "Élégance supérieure longue"
            complement_column["csv_name"] = "elegancesuperieurelongue"

            response = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-collision-validate"},
            )
            lot = client.get(f"/api/lots/{lot_id}").json()["lot"]

        self.assertEqual(accented["default_csv_name"], "b_elegance")
        self.assertEqual(response.status_code, 422)
        self.assertEqual(
            response.json()["error"]["code"], "SIRCOM_MAPPING_CSV_HEADER_COLLISION"
        )
        self.assertEqual(
            response.json()["error"]["details"]["warning_code"], "b_elegance"
        )
        self.assertEqual(step_status(lot, "mapping"), "bloque")
        self.assertIn(
            "SIRCOM_MAPPING_CSV_HEADER_COLLISION",
            {problem["code"] for problem in lot["problem_groups"]["bloquant"]["items"]},
        )

    def test_mapping_validation_requires_a_business_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures", ["valid_multi_tabs"]
            )
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot sans colonne métier",
                key="mapping-no-business",
            )
            mapping = client.get(f"/api/lots/{lot_id}/mapping").json()["mapping"]
            for column in mapping["columns"]:
                if column["system"] or column["logical_role"] == "id_dossier":
                    continue
                column["status"] = "supprime"

            response = client.post(
                f"/api/lots/{lot_id}/mapping/validate",
                json=mapping_submission(mapping),
                headers={"X-Idempotency-Key": "mapping-no-business-validate"},
            )

        self.assertEqual(response.status_code, 422)
        self.assertEqual(
            response.json()["error"]["code"], "SIRCOM_MAPPING_NO_BUSINESS_COLUMN"
        )


class MappingUiTest(unittest.TestCase):
    def test_home_ui_renders_mapping_unavailable_message(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            client = TestClient(create_app(settings))
            lot_id = client.post("/api/lots", json={"title": "Lot ancien"}).json()[
                "lot"
            ]["id"]

            with patch(
                "sircom2026.app.get_mapping_payload",
                side_effect=MappingError(
                    409,
                    "SIRCOM_MAPPING_SOURCE_HEADERS_MISSING",
                    "La liste structurée des colonnes est absente du diagnostic Excel.",
                ),
            ):
                response = client.get(f"/?lot_id={lot_id}&view=mapping")

        html = response.text
        self.assertEqual(response.status_code, 200)
        self.assertIn("Mapping indisponible", html)
        self.assertIn("SIRCOM_MAPPING_SOURCE_HEADERS_MISSING", html)
        self.assertIn("redéposer l&#39;Excel", html)

    def test_home_ui_renders_mapping_validation_screen_after_diagnostic(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            settings = make_settings(tmpdir)
            fixtures = create_synthetic_excels(
                tmpdir / "fixtures", ["valid_multi_tabs"]
            )
            client = TestClient(create_app(settings))
            lot_id = prepare_importable_lot(
                client,
                settings,
                fixtures["valid_multi_tabs"],
                title="Lot UI mapping",
                key="mapping-ui",
            )

            response = client.get(f"/lots/{lot_id}/excel?view=mapping")

        html = response.text
        self.assertEqual(response.status_code, 200)
        self.assertIn("Choisir les colonnes", html)
        self.assertIn('id="mapping-form"', html)
        self.assertEqual(html.count('id="breadcrumb-lot"'), 1)
        self.assertLess(
            html.index('id="breadcrumb-lot"'),
            html.index("Traitement Excel du lot Lot UI mapping"),
        )
        self.assertIn('id="header-workflow-menu"', html)
        self.assertIn("Parcours métier", html)
        self.assertIn("Traitement Excel", html)
        self.assertIn("Traitement images", html)
        self.assertIn("Export final", html)
        footer_html = html.split('<footer class="fr-footer" id="footer"', 1)[1]
        self.assertNotIn("Workflow d'orchestration", footer_html)
        self.assertNotIn("sircom-footer-workflow", footer_html)
        self.assertIn('id="lot-detail"', html)
        self.assertNotIn('id="overview-title"', html)
        self.assertIn("Colonne source", html)
        self.assertIn("Paramètres CSV", html)
        self.assertIn("sircom-mapping-columns-table", html)
        self.assertIn("fr-table--layout-fixed", html)
        self.assertIn("Nom CSV", html)
        self.assertIn("Rôle logique", html)
        self.assertIn("Dossiers", html)
        self.assertIn("Nom du produit", html)
        self.assertIn("d_nomdupro", html)
        self.assertIn("Décision à prendre sur le mapping", html)
        self.assertIn(
            'class="fr-link fr-icon-arrow-right-line fr-link--icon-right"', html
        )
        self.assertIn("Vérifier les colonnes", html)
        self.assertIn(
            'class="fr-link fr-icon-arrow-down-line fr-link--icon-right"', html
        )
        self.assertIn('data-mapping-bulk-action="select"', html)
        self.assertIn("Tout sélectionner", html)
        self.assertIn('data-mapping-bulk-action="deselect"', html)
        self.assertIn("Tout désélectionner", html)
        self.assertIn("Valider le mapping", html)
        self.assertIn("Sauvegarder le brouillon", html)
        self.assertIn('data-mapping-lot-id="', html)
        self.assertNotIn('href="#"', html)
        self.assertNotIn(str(tmpdir), html)


def mapping_submission(mapping: dict[str, Any]) -> dict[str, Any]:
    return {
        "structural_fingerprint": mapping["structural_fingerprint"],
        "columns": [
            {
                "id": column["id"],
                "status": column["status"],
                "csv_name": column["csv_name"],
                "logical_role": column["logical_role"],
                "suppression_reason": column["suppression_reason"],
            }
            for column in mapping["columns"]
        ],
    }


def find_column(
    mapping: dict[str, Any],
    sheet: str,
    letter: str,
    header: str,
) -> dict[str, Any]:
    for column in mapping["columns"]:
        if (
            column["source_sheet"] == sheet
            and column["source_column_letter"] == letter
            and column["source_header"] == header
        ):
            return column
    raise AssertionError(f"Missing column {sheet}!{letter}:{header}.")


def step_status(lot: dict[str, object], step_key: str) -> str:
    for step in lot["steps"]:
        if step["key"] == step_key:
            return str(step["status"])
    raise AssertionError(f"Missing step {step_key}.")


if __name__ == "__main__":
    unittest.main()
