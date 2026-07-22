from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts-2025" / "0-si-cellule-vide-na.py"


class Script2025FillEmptyNaTest(unittest.TestCase):
    def test_empty_cells_are_filled_without_losing_leading_zeroes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "Sircom.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Dossiers"
            sheet.append(["f_id", "code_postal", "telephone", "siret", "commentaire"])
            sheet.append(["00123", "01000", "0601020304", "01234567890123", None])
            sheet.append([None, "02000", "0700000000", "00000000000001", "ok"])
            workbook.save(workbook_path)
            workbook.close()

            result = subprocess.run(
                [sys.executable, str(SCRIPT)],
                cwd=tmp_path,
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            output_workbook = load_workbook(tmp_path / "Sircom_vide_na.xlsx")
            try:
                output_sheet = output_workbook["Dossiers"]
                values = {
                    "f_id": output_sheet["A2"].value,
                    "code_postal": output_sheet["B2"].value,
                    "telephone": output_sheet["C2"].value,
                    "siret": output_sheet["D2"].value,
                    "empty_comment": output_sheet["E2"].value,
                    "empty_id": output_sheet["A3"].value,
                    "second_siret": output_sheet["D3"].value,
                }
            finally:
                output_workbook.close()

        self.assertEqual(values["f_id"], "00123")
        self.assertEqual(values["code_postal"], "01000")
        self.assertEqual(values["telephone"], "0601020304")
        self.assertEqual(values["siret"], "01234567890123")
        self.assertEqual(values["empty_comment"], "#N/A")
        self.assertEqual(values["empty_id"], "#N/A")
        self.assertEqual(values["second_siret"], "00000000000001")


if __name__ == "__main__":
    unittest.main()
