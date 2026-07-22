from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts-2025" / "8-optimize_content_excel.py"


class Script2025OptimizeContentTest(unittest.TestCase):
    def test_crlf_cr_and_lf_become_single_indesign_br_markers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "7-add-pathimg.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["f_id", "description"])
            sheet.append(["ID-1", "Windows\r\nAncienne Mac\rUnix\nLinux"])
            workbook.save(workbook_path)
            workbook.close()

            result = subprocess.run(
                [sys.executable, str(SCRIPT)],
                cwd=tmp_path,
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            output_workbook = load_workbook(tmp_path / "8-optimize-content.xlsx")
            try:
                value = output_workbook.active["B2"].value
            finally:
                output_workbook.close()

        self.assertEqual(value, "Windows<br>Ancienne Mac<br>Unix<br>Linux")
        self.assertNotIn("<br><br>", value)


if __name__ == "__main__":
    unittest.main()
